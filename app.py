from flask import (
    Flask, render_template, request, redirect,
    url_for, send_file, flash, jsonify, abort, session
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os as _os
from dotenv import load_dotenv
load_dotenv()
from pathlib import Path
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

_BRT = ZoneInfo("America/Sao_Paulo")
from io import BytesIO
import json
import platform
import re
import subprocess
import unicodedata
import uuid

from services.gerar_contrato import gerar_docx, gerar_termo_quitacao, gerar_notificacao_avalista, gerar_notificacao_inadimplente, nome_arquivo_saida
from services.gerar_vistoria_entrada_saida import gerar_vistoria_entrada_saida, docx_para_pdf as _docx_para_pdf_es
from services import benchmarking_scraper

app = Flask(__name__)
app.secret_key = _os.environ.get("SECRET_KEY", "ativuz-secret-dev-2026")


@app.errorhandler(Exception)
def handle_any_error(e):
    import traceback; traceback.print_exc()
    return jsonify({"error": str(e)}), 500


# ── Template filters ──────────────────────────────────────────────────────────

@app.template_filter('brl')
def _fmt_brl(v):
    if v is None:
        return '—'
    neg = v < 0
    s = f"{abs(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"({s})" if neg else s


@app.template_filter('pct_fmt')
def _fmt_pct(v):
    if v is None:
        return '—'
    return f"{v * 100:.1f}%"


def _nh(s):
    """Normaliza string: minúsculas, sem acentos."""
    s = unicodedata.normalize("NFD", str(s or "").lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# ── Autenticação ──────────────────────────────────────────────────────────────

_ROTAS_PUBLICAS = {"login", "static", "admin_novo_usuario"}

@app.before_request
def verificar_login():
    if request.endpoint in _ROTAS_PUBLICAS:
        return
    if not session.get("usuario"):
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("usuario"):
        return redirect(url_for("dashboard"))
    erro = None
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        sb = _supabase()
        if not sb:
            erro = "Serviço indisponível. Tente novamente."
        else:
            try:
                from supabase import create_client
                url = _os.environ.get("SUPABASE_URL", "")
                key = _os.environ.get("SUPABASE_KEY", "")
                email = (nome if "@" in nome else f"{nome}@ativuz.com").lower()
                auth_client = create_client(url, key)
                res = auth_client.auth.sign_in_with_password({"email": email, "password": senha})
                session["usuario"] = nome
                return redirect(url_for("dashboard"))
            except Exception:
                erro = "Nome ou senha incorretos."
    return render_template("login.html", erro=erro)


@app.route("/logout")
def logout():
    sb = _supabase()
    if sb:
        try:
            sb.auth.sign_out()
        except Exception:
            pass
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/clientes")
def api_clientes():
    import openpyxl
    path = Path(__file__).parent / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"
    if not path.exists():
        return jsonify([])
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify([])
    # Mapeamento dinâmico por nome de coluna (case-insensitive, ignora acentos)
    import unicodedata
    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn")
    headers = [_norm(h) for h in rows[0]]
    def _col(name):
        n = _norm(name)
        return next((i for i, h in enumerate(headers) if n in h), None)
    i_nome    = _col("cliente")
    i_tel     = _col("telefone")
    i_ano     = _col("ano")
    i_chassi  = _col("chassi")
    i_cor     = _col("cor")
    i_marca   = _col("marca")
    i_modelo  = _col("modelo")
    i_placa   = _col("placa")
    i_end     = _col("endereco")
    i_motor   = _col("motor")
    def _v(row, i): return str(row[i] or "") if i is not None and i < len(row) else ""
    q = request.args.get("q", "").lower().strip()
    clientes = []
    for row in rows[1:]:
        if not (i_nome is not None and i_nome < len(row) and row[i_nome]):
            continue
        nome = str(row[i_nome])
        if q and q not in nome.lower():
            continue
        marca  = _v(row, i_marca)
        modelo = _v(row, i_modelo)
        clientes.append({
            "nome":         nome,
            "telefone":     _v(row, i_tel),
            "endereco":     _v(row, i_end),
            "veiculo":      f"{marca} {modelo}".strip(),
            "placa":        _v(row, i_placa),
            "cor":          _v(row, i_cor),
            "ano":          str(int(row[i_ano])) if i_ano is not None and i_ano < len(row) and row[i_ano] else "",
            "chassi":       _v(row, i_chassi),
            "numero_motor": _v(row, i_motor),
        })
    return jsonify(clientes[:30])


@app.route("/api/todos-telefones")
def api_todos_telefones():
    import openpyxl, unicodedata
    path = Path(__file__).parent / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"
    if not path.exists():
        return jsonify([])
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify([])
    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn")
    headers = [_norm(h) for h in rows[0]]
    def _col(name):
        n = _norm(name)
        return next((i for i, h in enumerate(headers) if n in h), None)
    i_nome = _col("cliente")
    i_tel  = _col("telefone")
    def _v(row, i): return str(row[i] or "").strip() if i is not None and i < len(row) else ""
    resultado = []
    vistos = set()
    for row in rows[1:]:
        if i_nome is None or i_nome >= len(row) or not row[i_nome]:
            continue
        nome = str(row[i_nome]).strip()
        fone = _v(row, i_tel)
        if not fone or not nome:
            continue
        if "segcomp" in _norm(nome):
            continue
        chave = (nome.lower(), fone)
        if chave in vistos:
            continue
        vistos.add(chave)
        resultado.append({"nome": nome, "telefone": fone})
    return jsonify(resultado)


@app.route("/api/asaas-parse", methods=["POST"])
def api_asaas_parse():
    import openpyxl, unicodedata
    f = request.files.get("arquivo")
    if not f:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    def _norm(s):
        s = unicodedata.normalize("NFD", str(s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn")

    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Localiza linha de cabeçalho real
    header_row = None
    for i, r in enumerate(rows):
        if r[0] == "Data":
            header_row = i
            break
    if header_row is None:
        return jsonify({"erro": "Formato de arquivo não reconhecido"}), 400

    # Extrai período
    periodo = ""
    for r in rows[:header_row]:
        for cell in r:
            s = str(cell or "")
            if "período" in s.lower() or "periodo" in s.lower():
                periodo = s
                break

    # Saldo inicial e final
    saldo_inicial = saldo_final = None
    for r in rows:
        desc = str(r[4] or "").strip()
        val  = r[6]
        if desc == "Saldo Inicial" and val is not None:
            saldo_inicial = float(val)
        if desc == "Saldo Final" and val is not None:
            saldo_final = float(val)

    transacoes = []
    for r in rows[header_row + 1:]:
        data   = str(r[0] or "").strip()
        tipo   = str(r[2] or "").strip()
        estorn = str(r[3] or "").strip()
        desc   = str(r[4] or "").strip()
        valor_raw = r[5]
        lancam = str(r[11] or "").strip()

        if not data or not desc or valor_raw is None:
            continue
        try:
            valor = float(valor_raw)
        except (TypeError, ValueError):
            continue

        desc_n = _norm(desc)
        abs_v  = abs(valor)

        # Classificação
        if estorn:
            categoria = "estorno"
        elif "cobrança recebida" in desc_n or "cobranca recebida" in desc_n:
            if abs_v >= 3000:
                categoria = "adesao"
            else:
                categoria = "aluguel"
        elif "luz divina" in desc_n:
            categoria = "repasse_investidor"
        elif "ativuz" in desc_n:
            categoria = "taxa_ativuz"
        elif "seguro" in desc_n:
            categoria = "seguro"
        elif "taxa" in desc_n or "notificacao" in desc_n or "notificação" in desc_n:
            categoria = "taxa_asaas"
        else:
            categoria = "outro"

        # Extrai nome do motorista (cobranças)
        motorista = ""
        if categoria in ("aluguel", "adesao"):
            import re as _re
            m = _re.search(r"fatura nr\.\s*\d+\s+(.+)$", desc, _re.IGNORECASE)
            motorista = m.group(1).strip() if m else ""

        # Placa do seguro
        placa_seguro = ""
        if categoria == "seguro":
            import re as _re
            m = _re.search(r"BYD\s+([A-Z0-9\-]+)", desc, _re.IGNORECASE)
            placa_seguro = m.group(1).upper().replace(" ", "") if m else ""

        transacoes.append({
            "data":          data,
            "tipo":          tipo,
            "descricao":     desc,
            "valor":         valor,
            "lancamento":    lancam,
            "categoria":     categoria,
            "motorista":     motorista,
            "placa_seguro":  placa_seguro,
        })

    # Totalizadores
    def _soma(cat):
        return sum(t["valor"] for t in transacoes if t["categoria"] == cat)

    totais = {
        "aluguel":            _soma("aluguel"),
        "adesao":             _soma("adesao"),
        "repasse_investidor": _soma("repasse_investidor"),
        "taxa_ativuz":        _soma("taxa_ativuz"),
        "seguro":             _soma("seguro"),
        "taxa_asaas":         _soma("taxa_asaas"),
        "estorno":            _soma("estorno"),
        "outro":              _soma("outro"),
    }

    return jsonify({
        "periodo":        periodo,
        "saldo_inicial":  saldo_inicial,
        "saldo_final":    saldo_final,
        "totais":         totais,
        "transacoes":     transacoes,
    })


@app.route("/admin/novo-usuario", methods=["GET", "POST"])
def admin_novo_usuario():
    token_correto = _os.environ.get("ADMIN_TOKEN", "")
    token = request.args.get("token", "")
    if not token_correto or token != token_correto:
        abort(403)
    mensagem = None
    erro = None
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        if not nome or not senha:
            erro = "Nome e senha são obrigatórios."
        else:
            sb = _supabase()
            if not sb:
                erro = "Supabase não configurado."
            else:
                try:
                    sb.table("usuarios").insert({
                        "nome": nome,
                        "senha_hash": generate_password_hash(senha),
                        "ativo": True,
                    }).execute()
                    mensagem = f"Usuário '{nome}' criado com sucesso!"
                except Exception as exc:
                    erro = f"Erro ao criar usuário: {exc}"
    return f"""
    <!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
    <title>Novo Usuário — Admin</title>
    <style>body{{font-family:Inter,sans-serif;background:#f0f2f7;display:flex;
    align-items:center;justify-content:center;min-height:100vh;}}
    .card{{background:#fff;border-radius:14px;padding:2rem;width:360px;
    box-shadow:0 4px 20px rgba(0,0,0,.1);}}
    h1{{font-size:1.1rem;margin-bottom:1.5rem;}}
    label{{font-size:.75rem;font-weight:600;text-transform:uppercase;
    letter-spacing:.05em;color:#475569;display:block;margin-bottom:.3rem;}}
    input{{width:100%;padding:.6rem .8rem;border:1.5px solid #e2e8f0;border-radius:8px;
    font-family:inherit;font-size:.9rem;margin-bottom:1rem;outline:none;}}
    input:focus{{border-color:#4361ee;}}
    button{{width:100%;padding:.7rem;background:#4361ee;color:#fff;border:none;
    border-radius:8px;font-weight:700;font-size:.9rem;cursor:pointer;}}
    .ok{{color:#166534;background:#f0fdf4;border:1px solid #bbf7d0;
    border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;}}
    .err{{color:#991b1b;background:#fef2f2;border:1px solid #fecaca;
    border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;}}
    </style></head><body><div class="card">
    <h1>Criar novo usuário</h1>
    {"<div class='ok'>"+mensagem+"</div>" if mensagem else ""}
    {"<div class='err'>"+erro+"</div>" if erro else ""}
    <form method="POST" action="?token={token}">
    <label>Nome</label><input name="nome" required>
    <label>Senha</label><input type="password" name="senha" required>
    <button>Criar usuário</button></form></div></body></html>
    """

# ── Supabase (opcional — só ativa se as env vars estiverem definidas) ─────────

_sb = None

def _supabase():
    global _sb
    if _sb is None:
        url = _os.environ.get("SUPABASE_URL", "")
        key = _os.environ.get("SUPABASE_KEY", "")
        if url and key:
            from supabase import create_client
            _sb = create_client(url, key)
    return _sb

UPLOAD_FOLDER = Path("uploads")
CONTRATOS_FOLDER = Path("contratos")
TEMP_FOLDER = Path("temp_preview")
DOCX_TEMPLATES = Path("docx_templates")

UPLOAD_FOLDER.mkdir(exist_ok=True)
CONTRATOS_FOLDER.mkdir(exist_ok=True)
TEMP_FOLDER.mkdir(exist_ok=True)
DOCX_TEMPLATES.mkdir(exist_ok=True)


# ── helpers ───────────────────────────────────────────────

def _converter_pdf(caminho_docx: str, caminho_pdf: str):
    """Converte .docx para PDF: Word no Windows, LibreOffice no Linux."""
    if platform.system() == "Windows":
        import pythoncom
        from docx2pdf import convert
        pythoncom.CoInitialize()
        try:
            convert(caminho_docx, caminho_pdf)
        finally:
            pythoncom.CoUninitialize()
    else:
        import tempfile, os, shutil
        docx_abs  = str(Path(caminho_docx).resolve())
        pdf_abs   = str(Path(caminho_pdf).resolve())
        if not Path(docx_abs).exists():
            raise FileNotFoundError(f"DOCX não encontrado: {docx_abs!r}")
        with tempfile.TemporaryDirectory() as work_dir:
            # copia para /tmp isolado — evita problemas de permissão/path no LO
            tmp_docx = Path(work_dir) / Path(docx_abs).name
            shutil.copy2(docx_abs, tmp_docx)
            env = {**os.environ, "HOME": work_dir}
            result = subprocess.run(
                [
                    "libreoffice",
                    "--headless", "--norestore", "--nofirststartwizard",
                    "--convert-to", "pdf",
                    "--outdir", work_dir,
                    str(tmp_docx),
                ],
                capture_output=True,
                env=env,
            )
            gerado = Path(work_dir) / (tmp_docx.stem + ".pdf")
            if not gerado.exists():
                stderr = result.stderr.decode(errors="replace") if result.stderr else ""
                stdout = result.stdout.decode(errors="replace") if result.stdout else ""
                raise RuntimeError(
                    f"LibreOffice (exit {result.returncode}) não gerou PDF. "
                    f"docx={str(tmp_docx)!r} stderr={stderr!r} stdout={stdout!r}"
                )
            shutil.copy2(gerado, pdf_abs)


def _slugify(texto: str) -> str:
    """Maiúsculas, sem acentos, espaços → underscore, sem caracteres especiais."""
    norm = unicodedata.normalize('NFD', texto)
    norm = ''.join(c for c in norm if unicodedata.category(c) != 'Mn')
    norm = norm.upper().strip()
    norm = re.sub(r'[^A-Z0-9\s]', '', norm)
    norm = re.sub(r'\s+', '_', norm)
    return norm or "SEM_NOME"


def detectar_tipo(filename: str):
    """Retorna o tipo do template com base no nome do arquivo."""
    norm = unicodedata.normalize('NFD', filename.lower())
    norm = ''.join(c for c in norm if unicodedata.category(c) != 'Mn')
    if 'quitacao' in norm:
        return 'quitacao'
    if 'locacao' in norm:
        return 'locacao'
    if 'notificacao' in norm and 'inadimplente' in norm:
        return 'inadimplente'
    if 'notificacao' in norm:
        return 'notificacao'
    return None


def get_templates():
    sb = _supabase()
    if sb:
        try:
            items = sb.storage.from_("documentos").list("templates") or []
            data_files = sorted([f for f in items if not f["name"].endswith(".json")], key=lambda x: x["name"])
            meta_map   = {f["name"]: True for f in items if f["name"].endswith(".json")}
            result = []
            for finfo in data_files:
                fname = finfo["name"]
                stem  = Path(fname).stem
                display_name = stem
                if f"{stem}.json" in meta_map:
                    try:
                        mb = sb.storage.from_("documentos").download(f"templates/{stem}.json")
                        display_name = json.loads(bytes(mb)).get("nome", stem)
                    except Exception:
                        pass
                size_kb = round((finfo.get("metadata") or {}).get("size", 0) / 1024, 1)
                updated = finfo.get("updated_at") or finfo.get("created_at") or ""
                try:
                    dt = datetime.fromisoformat(updated.replace("Z", "+00:00"))
                    data_fmt = dt.astimezone(_BRT).strftime("%d/%m/%Y %H:%M")
                except Exception:
                    data_fmt = ""
                result.append({"filename": fname, "nome": display_name,
                                "tamanho_kb": size_kb, "data": data_fmt})
            return result
        except Exception:
            import traceback; traceback.print_exc()
    # fallback local — une uploads/ e docx_templates/ sem duplicatas
    result = []
    seen = set()
    all_files = sorted(
        list(UPLOAD_FOLDER.glob("*.docx")) + list(UPLOAD_FOLDER.glob("*.xlsx")) +
        list(DOCX_TEMPLATES.glob("*.docx")) + list(DOCX_TEMPLATES.glob("*.xlsx")),
        key=lambda f: f.name,
    )
    for f in all_files:
        if f.name in seen:
            continue
        seen.add(f.name)
        # Busca metadados primeiro em uploads/, depois em docx_templates/
        meta_path = UPLOAD_FOLDER / f"{f.stem}.json"
        if not meta_path.exists():
            meta_path = DOCX_TEMPLATES / f"{f.stem}.json"
        display_name = f.stem
        if meta_path.exists():
            display_name = json.loads(meta_path.read_text(encoding="utf-8")).get("nome", f.stem)
        result.append({
            "filename": f.name,
            "nome": display_name,
            "tamanho_kb": round(f.stat().st_size / 1024, 1),
            "data": datetime.fromtimestamp(f.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        })
    return result


def _resolve_template(filename: str):
    """
    Retorna (caminho_local, nome_display, erro).
    Se Supabase disponível, baixa para arquivo temporário.
    Caller deve apagar o temp se caminho_local estiver em TEMP_FOLDER.
    """
    safe = secure_filename(filename)
    stem = Path(safe).stem
    sb   = _supabase()
    if sb:
        try:
            data = sb.storage.from_("documentos").download(f"templates/{safe}")
            if not data:
                return None, filename, "Template não encontrado no Storage."
            TEMP_FOLDER.mkdir(exist_ok=True)
            tmp = TEMP_FOLDER / f"tpl_{uuid.uuid4().hex}{Path(safe).suffix}"
            tmp.write_bytes(bytes(data))
            nome_display = stem
            try:
                mb = sb.storage.from_("documentos").download(f"templates/{stem}.json")
                nome_display = json.loads(bytes(mb)).get("nome", stem)
            except Exception:
                pass
            return str(tmp), nome_display, None
        except Exception as e:
            import traceback; traceback.print_exc()
    # fallback local — tenta uploads/ depois docx_templates/
    local = UPLOAD_FOLDER / safe
    if not local.exists():
        local = DOCX_TEMPLATES / safe
    if not local.exists():
        return None, filename, "Template não encontrado."
    nome_display = stem
    meta_local = UPLOAD_FOLDER / f"{stem}.json"
    if meta_local.exists():
        try:
            nome_display = json.loads(meta_local.read_text(encoding="utf-8")).get("nome", stem)
        except Exception:
            pass
    return str(local), nome_display, None


def _historico_append(locatario_nome: str, template: str, arquivo: str):
    sb = _supabase()
    if not sb:
        return
    try:
        sb.table("historico_docs").insert({
            "locatario_nome": locatario_nome,
            "template": template,
            "arquivo": arquivo,
            "data_hora": datetime.now(_BRT).strftime("%d/%m/%Y %H:%M"),
        }).execute()
    except Exception:
        import traceback; traceback.print_exc()


def _gerar_para_caminho(form, tipo, template_path_str, caminho_saida):
    """Gera o documento para caminho_saida. Retorna nome_pessoa."""
    if tipo == "locacao":
        campos = [
            "locatario_nome", "locatario_rg", "locatario_cpf",
            "locatario_endereco", "locatario_cep", "locatario_telefone",
            "avalista_nome", "avalista_cpf", "avalista_endereco", "avalista_telefone",
            "veiculo_descricao", "veiculo_marca", "veiculo_modelo", "veiculo_ano",
            "veiculo_motor", "veiculo_chassi", "veiculo_cor", "veiculo_placa",
            "contrato_inicio", "contrato_duracao", "valor_semanal",
            "data_dia", "data_mes", "data_ano",
            "testemunha1_nome", "testemunha1_rg", "testemunha1_cpf",
            "testemunha2_nome", "testemunha2_rg", "testemunha2_cpf",
        ]
        dados = {c: form.get(c, "") for c in campos}
        gerar_docx(dados, caminho_saida, template_path=template_path_str)
        return dados["locatario_nome"]

    elif tipo == "notificacao":
        avalista_nome = form.get("avalista_nome_notif", "")
        gerar_notificacao_avalista(
            avalista_nome  = avalista_nome,
            data_contrato  = form.get("data_contrato", ""),
            locatario_nome = form.get("locatario_nome_notif", ""),
            valor_debito   = float(form.get("valor_debito") or 0),
            caminho_saida  = caminho_saida,
            template_path  = template_path_str,
            avalista_cpf   = form.get("avalista_cpf_notif", ""),
        )
        return avalista_nome

    elif tipo == "inadimplente":
        locatario_nome_inad = form.get("locatario_nome_inad", "")
        gerar_notificacao_inadimplente(
            locatario_nome = locatario_nome_inad,
            data_contrato  = form.get("data_contrato_inad", ""),
            valor_debito   = float(form.get("valor_debito_inad") or 0),
            caminho_saida  = caminho_saida,
            template_path  = template_path_str,
        )
        return locatario_nome_inad

    else:  # quitacao
        def _f(campo): return float(form.get(campo) or 0)
        def _i(campo): return int(float(form.get(campo) or 0))
        devedor_nome = form.get("devedor_nome", "")
        gerar_termo_quitacao(
            devedor_nome          = devedor_nome,
            devedor_cpf           = form.get("devedor_cpf", ""),
            placa                 = form.get("placa", ""),
            mes_referencia_fipe   = form.get("mes_referencia_fipe", ""),
            valor_fipe            = _f("valor_fipe"),
            percentual_fipe       = _f("percentual_fipe"),
            meias_diarias         = _f("meias_diarias"),
            entrada               = _f("entrada"),
            num_parcelas_pagas    = _i("num_parcelas_pagas"),
            valor_parcela_paga    = _f("valor_parcela_paga"),
            num_parcelas_semanais = _i("num_parcelas_semanais"),
            valor_parcela_semanal = _f("valor_parcela_semanal"),
            data_primeira_parcela = form.get("data_primeira_parcela", ""),
            data_assinatura       = form.get("data_assinatura", ""),
            caminho_saida         = caminho_saida,
            template_path         = template_path_str,
        )
        return devedor_nome


# ── página 1 — Templates ──────────────────────────────────

@app.route("/")
def dashboard():
    sb = _supabase()
    total_contratos = 0
    total_vistorias = 0
    total_docs = 0
    valor_mensal = "—"
    contratos = []
    if sb:
        try:
            res = sb.table("contratos_locacao").select(
                "id, locatario_nome, veiculo_placa, veiculo_marca, veiculo_modelo, contrato_inicio, valor_semanal",
                count="exact"
            ).order("criado_em", desc=True).limit(5).execute()
            contratos = res.data or []
            total_contratos = res.count or len(contratos)
        except Exception:
            pass
        try:
            rv = sb.table("vistorias").select("id", count="exact").execute()
            total_vistorias = rv.count or 0
        except Exception:
            pass
    if sb:
        try:
            rd = sb.table("historico_docs").select("id", count="exact").eq("deletado", False).execute()
            total_docs = rd.count or 0
        except Exception:
            pass

    # ── Frota summary ──────────────────────────────────────
    frota_total = 0
    frota_valor_fipe = None
    try:
        veiculos_frota, _, _ = _ler_frota_dados()
        curr_key, curr_label, prev_key, _ = _frota_mes_atual()
        manual_frota = _frota_ler_manual()  # {placa: {mes_ref_label: {valor, ...}}}
        frota_total = len(veiculos_frota)
        total = 0.0
        any_val = False
        for v in veiculos_frota:
            placa = v.get("placa", "")
            meses_manual = manual_frota.get(placa) or {}
            mc = meses_manual.get(curr_label)
            if mc and mc.get("valor") is not None:
                total += float(mc["valor"])
                any_val = True
            elif v.get(curr_key) is not None:
                total += float(v[curr_key])
                any_val = True
            elif v.get(prev_key) is not None:
                total += float(v[prev_key])
                any_val = True
        if any_val:
            frota_valor_fipe = total
    except Exception:
        pass

    # ── Checklist: docs incompletas ─────────────────────────
    ck_pendentes_total = 0
    ck_pendentes_placas = []
    try:
        veiculos_ck, _ = _ler_veiculos()
        badge_data_ck = {}
        if sb:
            contratos_res = sb.table("checklist_contratos").select("id, contrato").execute()
            if contratos_res.data:
                ids_map = {r["id"]: r["contrato"] for r in contratos_res.data}
                itens_res = sb.table("checklist_itens").select("contrato_id, marcado").execute()
                for item in (itens_res.data or []):
                    cid  = item["contrato_id"]
                    cnum = ids_map.get(cid)
                    if cnum:
                        if cnum not in badge_data_ck:
                            badge_data_ck[cnum] = {"total": 0, "marcados": 0}
                        badge_data_ck[cnum]["total"] += 1
                        if item["marcado"]:
                            badge_data_ck[cnum]["marcados"] += 1
        for v in veiculos_ck:
            if not v.get("contrato"):
                continue
            bd = badge_data_ck.get(v["contrato"])
            if bd is None or (bd["total"] > 0 and bd["marcados"] < bd["total"]):
                ck_pendentes_placas.append(v["placa"])
        ck_pendentes_total = len(ck_pendentes_placas)
    except Exception:
        pass

    # ── Receita mensal real + contratos ativos ─────────────
    contratos_ativos = 0
    total_contratos_planilha = 0
    receita_mensal_real = None
    lista_contratos = []
    try:
        lista_contratos, _ = _ler_contratos()
        ativos_lst = [c for c in lista_contratos if c['situacao'] == 'EM ANDAMENTO']
        contratos_ativos = len(ativos_lst)
        total_contratos_planilha = len(lista_contratos)
        s = sum(c['valor_locacao'] for c in ativos_lst)
        if s > 0:
            receita_mensal_real = s
    except Exception:
        pass

    # ── Saldo de financiamentos ─────────────────────────────
    saldo_financiamentos = None
    fin_ativos = 0
    try:
        from math import ceil as _ceil
        hoje_fin = datetime.now(_BRT).date()
        rows_fin = (sb.table("financiamentos_contratos").select(
            "valor_parcela,data_vencimento,parcelas_total"
        ).execute().data or []) if sb else []
        saldo = 0.0
        for r in rows_fin:
            try:
                vcto = date.fromisoformat(str(r.get("data_vencimento", ""))[:10])
                dias = (vcto - hoje_fin).days
                restante = _ceil(dias / 30.44) if dias > 0 else 0
                if restante > 0:
                    fin_ativos += 1
                    saldo += restante * float(r["valor_parcela"])
            except Exception:
                continue
        if saldo > 0:
            saldo_financiamentos = saldo
    except Exception:
        pass

    # ── Carteira judicializada ──────────────────────────────
    jud_processos = 0
    jud_valor = None
    try:
        rows_jud = (sb.table("carteira_judicializada").select(
            "status,valor_atual"
        ).execute().data or []) if sb else []
        ativos_jud = [r for r in rows_jud if (r.get("status") or "").lower() != "perdido"]
        jud_processos = len(ativos_jud)
        v = sum(float(r.get("valor_atual") or 0) for r in ativos_jud)
        if v > 0:
            jud_valor = v
    except Exception:
        pass

    inad = _inad_summary()
    contratos_vencendo = _contratos_vencendo(dias_limite=60, contratos=lista_contratos or None)
    return render_template(
        "dashboard.html",
        active="dashboard",
        total_contratos=total_contratos_planilha or total_contratos,
        contratos_ativos=contratos_ativos,
        total_vistorias=total_vistorias,
        total_docs=total_docs,
        valor_mensal=receita_mensal_real,
        contratos=contratos,
        inad=inad,
        frota_total=frota_total,
        frota_valor_fipe=frota_valor_fipe,
        ck_pendentes_total=ck_pendentes_total,
        ck_pendentes_placas=ck_pendentes_placas,
        contratos_vencendo=contratos_vencendo,
        saldo_financiamentos=saldo_financiamentos,
        fin_ativos=fin_ativos,
        jud_processos=jud_processos,
        jud_valor=jud_valor,
    )


@app.route("/templates")
def pagina_templates():
    return render_template("templates.html", templates=get_templates(), active="templates")


@app.route("/upload", methods=["POST"])
def upload_template():
    nome = request.form.get("nome", "").strip()
    arquivo = request.files.get("arquivo")

    if not nome:
        flash("Informe um nome para o template.", "erro")
        return redirect(url_for("pagina_templates"))

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo .docx.", "erro")
        return redirect(url_for("pagina_templates"))

    if not arquivo.filename.lower().endswith((".docx", ".xlsx")):
        flash("Apenas arquivos .docx ou .xlsx são aceitos.", "erro")
        return redirect(url_for("pagina_templates"))

    uid  = uuid.uuid4().hex[:8]
    ext  = Path(secure_filename(arquivo.filename)).suffix.lower()
    safe_stem    = secure_filename(f"{nome}_{uid}")
    storage_path = f"templates/{safe_stem}{ext}"
    meta_path    = f"templates/{safe_stem}.json"

    sb = _supabase()
    if sb:
        try:
            file_bytes = arquivo.read()
            ct = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                  if ext == ".xlsx"
                  else "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
            sb.storage.from_("documentos").upload(storage_path, file_bytes,
                                                   {"content-type": ct, "upsert": "true"})
            sb.storage.from_("documentos").upload(
                meta_path,
                json.dumps({"nome": nome}, ensure_ascii=False).encode("utf-8"),
                {"content-type": "application/json", "upsert": "true"},
            )
        except Exception as e:
            flash(f"Erro ao salvar template: {e}", "erro")
            return redirect(url_for("pagina_templates"))
    else:
        dest = UPLOAD_FOLDER / f"{safe_stem}{ext}"
        arquivo.save(str(dest))
        (UPLOAD_FOLDER / f"{safe_stem}.json").write_text(
            json.dumps({"nome": nome}, ensure_ascii=False), encoding="utf-8"
        )

    flash(f'Template "{nome}" enviado com sucesso!', "ok")
    return redirect(url_for("pagina_templates"))


@app.route("/templates/excluir/<filename>", methods=["POST"])
def excluir_template(filename):
    safe = secure_filename(filename)
    stem = Path(safe).stem
    sb   = _supabase()
    if sb:
        try:
            sb.storage.from_("documentos").remove([f"templates/{safe}", f"templates/{stem}.json"])
        except Exception:
            import traceback; traceback.print_exc()
    else:
        for p in (UPLOAD_FOLDER / safe, UPLOAD_FOLDER / f"{stem}.json"):
            if p.exists():
                p.unlink()
    flash("Template excluído.", "ok")
    return redirect(url_for("pagina_templates"))


# ── página 2 — Gerar Contrato ─────────────────────────────

@app.route("/gerar")
def pagina_gerar():
    return render_template("gerar.html", templates=get_templates(), active="gerar")


@app.route("/gerar-contrato", methods=["POST"])
def gerar_contrato_route():
    template_filename = request.form.get("template", "")
    if not template_filename:
        flash("Selecione um template.", "erro")
        return redirect(url_for("pagina_gerar"))

    tipo = detectar_tipo(template_filename)
    if tipo is None:
        return jsonify({
            "error": "Template não reconhecido. Renomeie o arquivo com 'locacao', 'quitacao', 'notificacao' ou 'inadimplente' no nome."
        }), 400

    tpl_path, nome_template, tpl_erro = _resolve_template(template_filename)
    if tpl_erro:
        flash(tpl_erro, "erro")
        return redirect(url_for("pagina_gerar"))

    formato = request.form.get("formato", "docx")

    # ── Nome do arquivo de saída ──────────────────────────
    if tipo == "locacao":
        ano        = datetime.now().strftime("%Y")
        nome_saida = f"{ano}_{_slugify(request.form.get('veiculo_placa', ''))}_{_slugify(request.form.get('locatario_nome', ''))}.docx"
    elif tipo == "notificacao":
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"NOTIFICACAO_AVALISTA_{_slugify(request.form.get('avalista_nome_notif', ''))}_{data_slug}.docx"
    elif tipo == "inadimplente":
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"NOTIFICACAO_INADIMPLENTE_{_slugify(request.form.get('locatario_nome_inad', ''))}_{data_slug}.docx"
    else:  # quitacao
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"QUITACAO_DIVIDA_{_slugify(request.form.get('devedor_nome', ''))}_{data_slug}.docx"

    caminho_saida = str(CONTRATOS_FOLDER / nome_saida)

    # ── Gerar documento ───────────────────────────────────
    try:
        nome_pessoa = _gerar_para_caminho(request.form, tipo, tpl_path, caminho_saida)
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar contrato: {e}"}), 500
    finally:
        # remove temp se foi baixado do Storage
        if tpl_path and tpl_path.startswith(str(TEMP_FOLDER)):
            Path(tpl_path).unlink(missing_ok=True)
    try:
        _historico_append(nome_pessoa, nome_template, nome_saida)
    except Exception:
        pass

    # ── Download direto ────────────────────────────────────
    if formato == "pdf":
        nome_pdf    = nome_saida.replace(".docx", ".pdf")
        caminho_pdf = str(CONTRATOS_FOLDER / nome_pdf)
        try:
            _converter_pdf(caminho_saida, caminho_pdf)
            pdf_bytes = BytesIO(Path(caminho_pdf).read_bytes())
            return send_file(
                pdf_bytes,
                as_attachment=True,
                download_name=nome_pdf,
                mimetype="application/pdf",
            )
        except Exception as e:
            docx_url = url_for("download_contrato", filename=nome_saida)
            return jsonify({
                "error": f"Erro ao gerar PDF: {e}",
                "docx_url": docx_url,
            }), 422

    return send_file(
        caminho_saida,
        as_attachment=True,
        download_name=nome_saida,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route("/preview-contrato", methods=["POST"])
def preview_contrato():
    import mammoth

    template_filename = request.form.get("template", "")
    if not template_filename:
        return jsonify({"error": "Selecione um template."}), 400

    tipo = detectar_tipo(template_filename)
    if tipo is None:
        return jsonify({"error": "Template não reconhecido."}), 400
    if tipo == "vistoria":
        return jsonify({"error": "Pré-visualização não disponível para vistoria (formato .xlsx)."}), 400

    tpl_path, _, tpl_erro = _resolve_template(template_filename)
    if tpl_erro:
        return jsonify({"error": tpl_erro}), 400

    temp_id = uuid.uuid4().hex
    caminho_temp = str(TEMP_FOLDER / f"{temp_id}.docx")

    try:
        _gerar_para_caminho(request.form, tipo, tpl_path, caminho_temp)
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar pré-visualização: {e}"}), 500
    finally:
        if tpl_path and tpl_path.startswith(str(TEMP_FOLDER)) and tpl_path != caminho_temp:
            Path(tpl_path).unlink(missing_ok=True)

    try:
        with open(caminho_temp, "rb") as f:
            result = mammoth.convert_to_html(f)
        html = result.value
    except Exception as e:
        Path(caminho_temp).unlink(missing_ok=True)
        return jsonify({"error": f"Erro ao converter para HTML: {e}"}), 500

    return jsonify({"html": html, "temp_id": temp_id})


@app.route("/cleanup-temp/<temp_id>", methods=["POST"])
def cleanup_temp(temp_id):
    if not re.match(r'^[0-9a-f]{32}$', temp_id):
        abort(400)
    caminho = TEMP_FOLDER / f"{temp_id}.docx"
    if caminho.exists():
        caminho.unlink()
    return jsonify({"ok": True})


# ── página 3 — Histórico ──────────────────────────────────

@app.route("/historico")
def pagina_historico():
    sb = _supabase()
    historico = []
    if sb:
        try:
            res = sb.table("historico_docs").select("*") \
                .eq("deletado", False) \
                .order("criado_em", desc=True).execute()
            historico = res.data or []
        except Exception:
            pass
    return render_template("historico.html", historico=historico, active="historico")


@app.route("/historico/download/<path:filename>")
def download_contrato(filename):
    caminho = (CONTRATOS_FOLDER / filename).resolve()
    if not str(caminho).startswith(str(CONTRATOS_FOLDER.resolve())):
        abort(400)
    if not caminho.exists():
        flash("Arquivo não encontrado.", "erro")
        return redirect(url_for("pagina_historico"))
    ext = Path(filename).suffix.lower()
    mime = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if ext == ".xlsx"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    return send_file(
        str(caminho),
        as_attachment=True,
        download_name=Path(filename).name,
        mimetype=mime,
    )


@app.route("/historico/download-pdf/<path:filename>")
def download_contrato_pdf(filename):
    caminho_docx = (CONTRATOS_FOLDER / filename).resolve()
    if not str(caminho_docx).startswith(str(CONTRATOS_FOLDER.resolve())):
        abort(400)
    if not caminho_docx.exists():
        return jsonify({"error": "Arquivo não encontrado."}), 404

    nome_pdf    = Path(filename).stem + ".pdf"
    caminho_pdf = CONTRATOS_FOLDER / nome_pdf
    try:
        _converter_pdf(str(caminho_docx), str(caminho_pdf))
        pdf_bytes = BytesIO(Path(caminho_pdf).read_bytes())
        return send_file(
            pdf_bytes,
            as_attachment=True,
            download_name=nome_pdf,
            mimetype="application/pdf",
        )
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar PDF: {e}"}), 422


@app.route("/historico/excluir/<entry_id>", methods=["POST"])
def excluir_contrato(entry_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("historico_docs").update({"deletado": True}).eq("id", entry_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/exportar-excel")
def exportar_historico_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    sb = _supabase()
    historico = []
    if sb:
        try:
            res = sb.table("historico_docs").select("*") \
                .eq("deletado", False) \
                .order("criado_em", desc=True).execute()
            historico = res.data or []
        except Exception:
            pass
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"

    cabecalho = ["Locatário", "Template", "Data / Hora", "Nome do Arquivo"]
    ws.append(cabecalho)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    for item in historico:
        ws.append([
            item.get("locatario_nome", ""),
            item.get("template", ""),
            item.get("data_hora", ""),
            item.get("arquivo", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    data_hoje = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"HISTORICO_ATIVUZ_{data_hoje}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Contrato de Locação — helpers e rotas ────────────────────────────────────

CONTRATO_LOCACAO_TEMPLATE = DOCX_TEMPLATES / "CONTRATO DE LOCAÇÃO EDITADO.docx"


def _salvar_contrato_locacao(insert: dict, caminho_docx: str, storage_path: str, edit_id: str = None):
    """INSERT no Supabase (main thread) + upload do arquivo (background).
    Retorna None em sucesso ou str com mensagem de erro."""
    import threading, traceback as _tb
    sb = _supabase()
    if not sb:
        return "Supabase não configurado."
    try:
        sb.table("contratos_locacao").insert(insert).execute()
    except Exception as e:
        _tb.print_exc()
        return str(e)

    # Só remove o registro antigo APÓS o INSERT ter sido bem-sucedido
    _old_path = None
    if edit_id:
        try:
            old = sb.table("contratos_locacao").select("arquivo_path").eq("id", edit_id).single().execute()
            _old_path = (old.data or {}).get("arquivo_path")
            sb.table("contratos_locacao").delete().eq("id", edit_id).execute()
        except Exception:
            _tb.print_exc()

    _docx_bytes = Path(caminho_docx).read_bytes()
    _sp = storage_path
    _op = _old_path

    def _bg():
        try:
            sb2 = _supabase()
            if not sb2:
                return
            try:
                sb2.storage.from_("documentos").upload(
                    _sp, _docx_bytes,
                    {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                     "upsert": "true"},
                )
            except Exception:
                _tb.print_exc()
            if _op and _op != _sp:
                try:
                    sb2.storage.from_("documentos").remove([_op])
                except Exception:
                    pass
        except Exception:
            _tb.print_exc()

    threading.Thread(target=_bg, daemon=True).start()
    return None

_MESES_PT = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]


@app.route("/contrato-locacao")
def pagina_contrato_locacao():
    agora = datetime.now(_BRT)
    defaults = {
        "data_dia": agora.strftime("%d"),
        "data_mes": _MESES_PT[agora.month - 1],
        "data_ano": agora.strftime("%Y"),
    }
    return render_template("contrato_locacao.html", defaults=defaults, active="contrato_locacao")


@app.route("/contrato-locacao/gerar", methods=["POST"])
def gerar_contrato_locacao_route():
    """Usado pelo fluxo de edição a partir de historico_contratos."""
    if not CONTRATO_LOCACAO_TEMPLATE.exists():
        return jsonify({"error": "Template não encontrado em docx_templates/."}), 404

    campos = [
        "locatario_nome", "locatario_rg", "locatario_cpf",
        "locatario_endereco", "locatario_cep", "locatario_telefone",
        "avalista_nome", "avalista_cpf", "avalista_endereco", "avalista_telefone",
        "veiculo_descricao", "veiculo_marca", "veiculo_modelo", "veiculo_ano",
        "veiculo_motor", "veiculo_chassi", "veiculo_cor", "veiculo_placa",
        "contrato_inicio", "contrato_duracao", "valor_semanal",
        "caucao_valor", "caucao_extenso",
        "data_dia", "data_mes", "data_ano",
        "testemunha1_nome", "testemunha1_rg", "testemunha1_cpf",
        "testemunha2_nome", "testemunha2_rg", "testemunha2_cpf",
    ]
    dados   = {c: request.form.get(c, "") for c in campos}
    edit_id = request.form.get("edit_id", "").strip()

    placa_slug    = _slugify(dados.get("veiculo_placa") or "PLACA")
    nome_slug     = _slugify((dados.get("locatario_nome") or "LOCATARIO").split()[0])
    data_slug     = datetime.now(_BRT).strftime("%d.%m.%Y")
    nome_docx     = f"CONTRATO_LOCACAO_{placa_slug}_{nome_slug}_{data_slug}.docx"
    caminho_saida = str(CONTRATOS_FOLDER / nome_docx)

    try:
        gerar_docx(dados, caminho_saida, template_path=str(CONTRATO_LOCACAO_TEMPLATE))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao gerar contrato: {e}"}), 500

    _storage_path = f"contratos/{nome_docx}"
    _insert = {**dados, "arquivo_path": _storage_path}
    # Campos que não existem na tabela contratos_locacao do Supabase
    for _campo_extra in ("caucao_extenso", "caucao_valor"):
        _insert.pop(_campo_extra, None)
    _err = _salvar_contrato_locacao(_insert, caminho_saida, _storage_path, edit_id=edit_id or None)
    if _err:
        return jsonify({"error": f"Erro ao salvar no banco de dados: {_err}"}), 500

    return jsonify({"redirect_url": url_for("historico_contratos")})


@app.route("/historico/contratos")
def historico_contratos():
    sb = _supabase()
    contratos = []
    erro = None
    if sb:
        try:
            res = sb.table("contratos_locacao").select(
                "id, locatario_nome, locatario_cpf, veiculo_placa, veiculo_marca, "
                "veiculo_modelo, contrato_inicio, valor_semanal, arquivo_path, criado_em"
            ).neq("deletado", True).order("criado_em", desc=True).execute()
            contratos = res.data or []
        except Exception as e:
            erro = str(e)
    else:
        erro = "Supabase não configurado."
    return render_template("historico_contratos.html", contratos=contratos, erro=erro,
                           active="hist_contratos")


@app.route("/historico/contratos/<contrato_id>/excluir", methods=["POST"])
def excluir_contrato_locacao(contrato_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("contratos_locacao").update({"deletado": True}).eq("id", contrato_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/contratos/download/<contrato_id>")
def download_contrato_locacao_docx(contrato_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("contratos_locacao").select("arquivo_path").eq("id", contrato_id).single().execute()
        path = res.data["arquivo_path"]
        signed = sb.storage.from_("documentos").create_signed_url(path, 60)
        return redirect(signed["signedURL"])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/historico/contratos/download/<contrato_id>/pdf")
def download_contrato_locacao_pdf(contrato_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("contratos_locacao").select("arquivo_path").eq("id", contrato_id).single().execute()
        docx_path = res.data["arquivo_path"]
        docx_bytes = sb.storage.from_("documentos").download(docx_path)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if not isinstance(docx_bytes, (bytes, bytearray)):
        docx_bytes = getattr(docx_bytes, 'content', None) or bytes(docx_bytes)

    tmp_docx = TEMP_FOLDER / f"{uuid.uuid4().hex}.docx"
    tmp_pdf  = tmp_docx.with_suffix(".pdf")
    TEMP_FOLDER.mkdir(exist_ok=True)
    try:
        tmp_docx.write_bytes(docx_bytes)
        _converter_pdf(str(tmp_docx), str(tmp_pdf))
        if not tmp_pdf.exists():
            raise FileNotFoundError(f"LibreOffice não gerou o PDF")
        pdf_bytes = tmp_pdf.read_bytes()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao gerar PDF: {e}"}), 422
    finally:
        tmp_docx.unlink(missing_ok=True)
        tmp_pdf.unlink(missing_ok=True)

    nome_pdf = Path(docx_path).stem + ".pdf"
    return send_file(BytesIO(pdf_bytes), as_attachment=True,
                     download_name=nome_pdf, mimetype="application/pdf")


@app.route("/historico/contratos/<contrato_id>/editar")
def editar_contrato_locacao(contrato_id):
    sb = _supabase()
    if not sb:
        flash("Supabase não configurado.", "erro")
        return redirect(url_for("historico_contratos"))
    try:
        res = sb.table("contratos_locacao").select("*").eq("id", contrato_id).single().execute()
        contrato = res.data
    except Exception as e:
        flash(f"Erro ao buscar contrato: {e}", "erro")
        return redirect(url_for("historico_contratos"))
    agora = datetime.now(_BRT)
    defaults = {
        "data_dia": contrato.get("data_dia") or agora.strftime("%d"),
        "data_mes": contrato.get("data_mes") or _MESES_PT[agora.month - 1],
        "data_ano": contrato.get("data_ano") or agora.strftime("%Y"),
    }
    return render_template("contrato_locacao.html", contrato=contrato,
                           edit_id=contrato_id, defaults=defaults,
                           active="hist_contratos")


# ── Vistoria de Entrega ───────────────────────────────────────────────────────

VISTORIA_ES_TEMPLATE = DOCX_TEMPLATES / "VISTORIA_ENTRADA_SAIDA_TEMPLATE.docx"


@app.route("/vistoria", methods=["GET"])
def pagina_vistoria():
    return render_template("vistoria.html", active="vistoria", vistoria=None,
                           contrato_id=None, edit_id=None, acessorios={},
                           usuario=session.get("usuario", "").split("@")[0].split(".")[0].upper())


@app.route("/vistoria/<contrato_id>", methods=["GET"])
def pagina_vistoria_contrato(contrato_id):
    """Exibe o formulário de vistoria no estado correto (entrada / saida / completa)."""
    sb = _supabase()
    vistoria = None
    if sb:
        try:
            res = (sb.table("vistorias")
                     .select("*")
                     .eq("contrato_id", contrato_id)
                     .order("criado_em", desc=True)
                     .limit(1)
                     .execute())
            if res.data:
                vistoria = res.data[0]
        except Exception:
            import traceback; traceback.print_exc()
    return render_template("vistoria.html", active="vistoria",
                           vistoria=vistoria, contrato_id=contrato_id,
                           edit_id=None,
                           acessorios=(vistoria or {}).get("acessorios") or {},
                           usuario=session.get("usuario", "").split("@")[0].split(".")[0].upper())




@app.route("/vistoria/gerar", methods=["POST"])
def gerar_vistoria_route():
    foto_path = None
    try:
        return _gerar_vistoria_impl()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro interno: {e}"}), 500
    finally:
        if foto_path:
            Path(foto_path).unlink(missing_ok=True)


def _gerar_vistoria_impl():
    import threading, traceback as _tb
    agora = datetime.now(_BRT)
    etapa = request.form.get("etapa", "").strip()  # "entrada" | "saida" | "" (legado)

    _ANGULOS_FOTO = [
        "frontal", "traseira", "lateral_dir", "lateral_esq",
        "painel", "hodometro", "estepe", "teto",
        "motor", "mala", "dano_1", "dano_2",
    ]

    _CHAVES_ACC = [
        "acc_calotas", "acc_buzina", "acc_doc_crlv", "acc_triangulo", "acc_antena",
        "acc_sensor_re", "acc_som", "acc_tapetes", "acc_limpadores", "acc_chave_roda",
        "acc_vidros_eletricos", "acc_oleo_motor", "acc_alarme", "acc_lampadas", "acc_macaco",
        "acc_estepe", "acc_gnv", "acc_agua", "acc_borr_psg_dir", "acc_borr_mtr_dir",
        "acc_asa_dd", "acc_asa_td", "acc_tapete_mala", "acc_tampa_parachoque",
        "acc_borr_psg_tras", "acc_borr_mtr_tras", "acc_asa_de", "acc_asa_te",
        "acc_bagagito", "acc_lingueta",
    ]

    def _salvar_foto(file_storage):
        if not file_storage or not file_storage.filename:
            return None
        ext = Path(secure_filename(file_storage.filename)).suffix.lower()
        if ext not in ('.jpg', '.jpeg', '.png'):
            return None
        p = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
        file_storage.save(str(p))
        return str(p)

    def _upload_bg(storage_path, docx_bytes, old_storage_path=None):
        _sp, _db, _ost = storage_path, docx_bytes, old_storage_path
        def _bg():
            try:
                sb2 = _supabase()
                if not sb2:
                    return
                try:
                    sb2.storage.from_("documentos").upload(
                        _sp, _db,
                        {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                         "upsert": "true"},
                    )
                except Exception:
                    _tb.print_exc()
                if _ost and _ost != _sp:
                    try:
                        sb2.storage.from_("documentos").remove([_ost])
                    except Exception:
                        pass
            except Exception:
                _tb.print_exc()
        threading.Thread(target=_bg, daemon=True).start()

    dados_fixos = {
        "contrato_id":      request.form.get("contrato_id", "").strip(),
        "cliente_nome":     request.form.get("cliente_nome", ""),
        "cliente_telefone": request.form.get("cliente_telefone", ""),
        "cliente_endereco": request.form.get("cliente_endereco", ""),
        "preenchido_por":   request.form.get("preenchido_por", ""),
        "veiculo":          request.form.get("veiculo", "").strip(),
        "placa":            request.form.get("placa", "").upper().strip(),
        "cor":              request.form.get("cor", ""),
        "ano":              request.form.get("ano", ""),
        "chassi":           request.form.get("chassi", ""),
        "numero_motor":     request.form.get("numero_motor", ""),
    }
    contrato_id = dados_fixos["contrato_id"]
    placa_slug  = _slugify(dados_fixos["placa"] or "PLACA")

    # ─────────────────────────────────────────────────────────────────────────
    # ETAPA ENTRADA  (cliente retira o carro)
    # ─────────────────────────────────────────────────────────────────────────
    if etapa == "entrada":
        fotos_entrada = {}
        for angulo in _ANGULOS_FOTO:
            p = _salvar_foto(request.files.get(f"foto_entrada_{angulo}"))
            if p:
                fotos_entrada[angulo] = p

        dados = {
            **dados_fixos,
            "data_entrada":        agora.strftime("%d/%m/%Y %H:%M"),
            "hodometro_entrada":   request.form.get("hodometro_entrada", ""),
            "combustivel_entrada": request.form.get("combustivel_entrada", ""),
            "obs_entrada":         request.form.get("obs_entrada", ""),
            "sintomas_entrada":    request.form.get("sintomas_entrada", ""),
            "responsavel_entrada": dados_fixos["preenchido_por"],
            "acessorios_entrada":  {k: request.form.get(f"{k}_entrada", "") for k in _CHAVES_ACC},
            "fotos_entrada":       fotos_entrada,
        }

        data_slug    = agora.strftime("%d.%m.%Y")
        nome_docx    = f"VISTORIA_{placa_slug}_{data_slug}.docx"
        caminho_docx = str(CONTRATOS_FOLDER / nome_docx)

        # Lê bytes das fotos ANTES do finally apagar os arquivos temporários
        pasta_fotos   = f"vistorias/fotos/{placa_slug}_{data_slug}"
        foto_s_paths  = {}   # {angulo: storage_path}
        foto_s_bytes  = {}   # {storage_path: bytes} — para upload em background
        for angulo, local_p in fotos_entrada.items():
            ext = Path(local_p).suffix.lower() or ".jpg"
            s_path = f"{pasta_fotos}/{angulo}{ext}"
            foto_s_paths[angulo] = s_path
            try:
                foto_s_bytes[s_path] = Path(local_p).read_bytes()
            except Exception:
                pass

        try:
            resumo = gerar_vistoria_entrada_saida(
                dados,
                caminho_saida=caminho_docx,
                template_path=str(VISTORIA_ES_TEMPLATE),
            )
        except Exception as e:
            _tb.print_exc()
            return jsonify({"error": f"Erro ao gerar vistoria (entrada): {e}"}), 500
        finally:
            for p in fotos_entrada.values():
                Path(p).unlink(missing_ok=True)

        # fotos_entrada no banco: ["angulo:storage_path", ...] (coluna text[])
        fotos_entrada_db = [f"{ang}:{pth}" for ang, pth in foto_s_paths.items()]

        _storage_path = f"vistorias/{nome_docx}"
        sb = _supabase()
        if sb:
            try:
                sb.table("vistorias").insert({
                    "contrato_id":          contrato_id or None,
                    "cliente":              dados["cliente_nome"],
                    "telefone":             dados["cliente_telefone"],
                    "endereco":             dados["cliente_endereco"],
                    "preenchido_por":       dados["preenchido_por"],
                    "veiculo":              dados["veiculo"],
                    "placa":                dados["placa"],
                    "cor":                  dados["cor"],
                    "ano":                  dados["ano"],
                    "chassi":               dados["chassi"],
                    "numero_motor":         dados["numero_motor"],
                    "data_hora":            agora.strftime("%d/%m/%Y %H:%M"),
                    "data_entrada":         dados["data_entrada"],
                    "hodometro_entrada":    dados["hodometro_entrada"],
                    "combustivel_entrada":  dados["combustivel_entrada"],
                    "obs_entrada":          dados["obs_entrada"],
                    "sintomas_entrada":     dados["sintomas_entrada"],
                    "responsavel_entrada":  dados["responsavel_entrada"],
                    "acessorios_entrada":   dados["acessorios_entrada"],
                    "fotos_entrada":        fotos_entrada_db,
                    "status":               resumo["status"],
                    "arquivo_entrada_path": _storage_path,
                    "arquivo_path":         _storage_path,
                }).execute()
            except Exception:
                _tb.print_exc()
            _upload_bg(_storage_path, Path(caminho_docx).read_bytes())
            # Upload individual das fotos em background
            if foto_s_bytes:
                _fsb = foto_s_bytes
                def _upload_fotos_bg():
                    try:
                        sb2 = _supabase()
                        if not sb2:
                            return
                        for s_path, data in _fsb.items():
                            ext2 = Path(s_path).suffix.lower()
                            ct = "image/png" if ext2 == ".png" else "image/jpeg"
                            try:
                                sb2.storage.from_("documentos").upload(
                                    s_path, data, {"content-type": ct, "upsert": "true"})
                            except Exception:
                                _tb.print_exc()
                    except Exception:
                        _tb.print_exc()
                threading.Thread(target=_upload_fotos_bg, daemon=True).start()

        try:
            _historico_append(dados["cliente_nome"], "VISTORIA", nome_docx)
        except Exception:
            _tb.print_exc()

        return jsonify({"redirect_url": url_for("historico_vistorias")})

    # ─────────────────────────────────────────────────────────────────────────
    # ETAPA SAÍDA  (cliente devolve o carro)
    # ─────────────────────────────────────────────────────────────────────────
    if etapa == "saida":
        vistoria_id = request.form.get("vistoria_id", "").strip()
        sb = _supabase()
        registro = {}
        caminho_docx_anterior = None

        if sb:
            try:
                if vistoria_id:
                    res = sb.table("vistorias").select("*").eq("id", vistoria_id).execute()
                else:
                    res = (sb.table("vistorias").select("*")
                             .eq("contrato_id", contrato_id)
                             .order("criado_em", desc=True)
                             .limit(1)
                             .execute())
                if res.data:
                    registro = res.data[0]
                    vistoria_id = registro.get("id", vistoria_id)
                    caminho_docx_anterior = (registro.get("arquivo_entrada_path")
                                             or registro.get("arquivo_path"))
            except Exception:
                _tb.print_exc()

        fotos_saida = {}
        for angulo in _ANGULOS_FOTO:
            p = _salvar_foto(request.files.get(f"foto_saida_{angulo}"))
            if p:
                fotos_saida[angulo] = p

        # Recupera fotos de entrada do Storage (guardadas na etapa anterior)
        # Formato no banco: ["angulo:storage_path", ...]
        fotos_entrada_recuperadas = {}   # {angulo: caminho_temp_local}
        if sb:
            for item in (registro.get("fotos_entrada") or []):
                try:
                    angulo, s_path = item.split(":", 1)
                    data_foto = sb.storage.from_("documentos").download(s_path)
                    if data_foto:
                        ext_rec = Path(s_path).suffix.lower() or ".jpg"
                        tmp_rec = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext_rec}"
                        tmp_rec.write_bytes(bytes(data_foto))
                        fotos_entrada_recuperadas[angulo] = str(tmp_rec)
                except Exception:
                    _tb.print_exc()

        dados = {
            "contrato_id":      contrato_id or registro.get("contrato_id", ""),
            "cliente_nome":     registro.get("cliente", dados_fixos["cliente_nome"]),
            "cliente_telefone": registro.get("telefone", dados_fixos["cliente_telefone"]),
            "cliente_endereco": registro.get("endereco", dados_fixos["cliente_endereco"]),
            "preenchido_por":   registro.get("preenchido_por", dados_fixos["preenchido_por"]),
            "veiculo":          registro.get("veiculo", dados_fixos["veiculo"]),
            "placa":            registro.get("placa", dados_fixos["placa"]),
            "cor":              registro.get("cor", dados_fixos["cor"]),
            "ano":              registro.get("ano", dados_fixos["ano"]),
            "chassi":           registro.get("chassi", dados_fixos["chassi"]),
            "numero_motor":     registro.get("numero_motor", dados_fixos["numero_motor"]),
            # Entrega — do registro existente
            "data_entrada":        registro.get("data_entrada", ""),
            "hodometro_entrada":   registro.get("hodometro_entrada", ""),
            "combustivel_entrada": registro.get("combustivel_entrada", ""),
            "obs_entrada":         registro.get("obs_entrada", ""),
            "sintomas_entrada":    registro.get("sintomas_entrada", ""),
            "responsavel_entrada": registro.get("responsavel_entrada", ""),
            "acessorios_entrada":  registro.get("acessorios_entrada") or {},
            "fotos_entrada":       fotos_entrada_recuperadas,
            # Devolução — do form
            "data_saida":        agora.strftime("%d/%m/%Y %H:%M"),
            "hodometro_saida":   request.form.get("hodometro_saida", ""),
            "combustivel_saida": request.form.get("combustivel_saida", ""),
            "obs_saida":         request.form.get("obs_saida", ""),
            "sintomas_saida":    request.form.get("sintomas_saida", ""),
            "responsavel_saida": dados_fixos["preenchido_por"],
            "acessorios_saida":  {k: request.form.get(f"{k}_saida", "") for k in _CHAVES_ACC},
            "fotos_saida":       fotos_saida,
        }

        # Lê bytes das fotos de SAÍDA antes do finally apagar os temporários
        placa_saida      = registro.get("placa") or dados_fixos.get("placa") or "PLACA"
        data_slug_saida  = agora.strftime("%d.%m.%Y")
        pasta_fotos_saida = f"vistorias/fotos/{_slugify(placa_saida)}_{data_slug_saida}_saida"
        foto_saida_s_paths = {}   # {angulo: storage_path}
        foto_saida_s_bytes = {}   # {storage_path: bytes}
        for angulo, local_p in fotos_saida.items():
            ext = Path(local_p).suffix.lower() or ".jpg"
            s_path = f"{pasta_fotos_saida}/{angulo}{ext}"
            foto_saida_s_paths[angulo] = s_path
            try:
                foto_saida_s_bytes[s_path] = Path(local_p).read_bytes()
            except Exception:
                pass

        if caminho_docx_anterior:
            nome_docx    = Path(caminho_docx_anterior).name
            caminho_docx = str(CONTRATOS_FOLDER / nome_docx)
        else:
            nome_docx    = f"VISTORIA_{_slugify(placa_saida)}_{data_slug_saida}.docx"
            caminho_docx = str(CONTRATOS_FOLDER / nome_docx)

        try:
            resumo = gerar_vistoria_entrada_saida(
                dados,
                caminho_saida=caminho_docx,
                template_path=str(VISTORIA_ES_TEMPLATE),
            )
        except Exception as e:
            _tb.print_exc()
            return jsonify({"error": f"Erro ao gerar vistoria (saida): {e}"}), 500
        finally:
            for p in fotos_saida.values():
                Path(p).unlink(missing_ok=True)
            for p in fotos_entrada_recuperadas.values():
                Path(p).unlink(missing_ok=True)

        fotos_saida_db = [f"{ang}:{pth}" for ang, pth in foto_saida_s_paths.items()]

        _storage_path = f"vistorias/{nome_docx}"
        if sb and vistoria_id:
            try:
                sb.table("vistorias").update({
                    "data_saida":            dados["data_saida"],
                    "hodometro_saida":       dados["hodometro_saida"],
                    "combustivel_saida":     dados["combustivel_saida"],
                    "obs_saida":             dados["obs_saida"],
                    "sintomas_saida":        dados["sintomas_saida"],
                    "responsavel_saida":     dados["responsavel_saida"],
                    "acessorios_saida":      dados["acessorios_saida"],
                    "fotos_saida":           fotos_saida_db,
                    "status":                resumo["status"],
                    "divergencias":          [list(d) for d in resumo["divergencias"]],
                    "arquivo_completo_path": _storage_path,
                    "arquivo_path":          _storage_path,
                }).eq("id", vistoria_id).execute()
            except Exception:
                _tb.print_exc()
            _upload_bg(_storage_path, Path(caminho_docx).read_bytes())
            if foto_saida_s_bytes:
                _fsb2 = foto_saida_s_bytes
                def _upload_fotos_saida_bg():
                    try:
                        sb2 = _supabase()
                        if not sb2:
                            return
                        for s_path, data in _fsb2.items():
                            ext2 = Path(s_path).suffix.lower()
                            ct = "image/png" if ext2 == ".png" else "image/jpeg"
                            try:
                                sb2.storage.from_("documentos").upload(
                                    s_path, data, {"content-type": ct, "upsert": "true"})
                            except Exception:
                                _tb.print_exc()
                    except Exception:
                        _tb.print_exc()
                threading.Thread(target=_upload_fotos_saida_bg, daemon=True).start()

        try:
            _historico_append(dados["cliente_nome"], "VISTORIA", nome_docx)
        except Exception:
            _tb.print_exc()

        return jsonify({"redirect_url": url_for("historico_vistorias")})

    return jsonify({"error": "Etapa de vistoria inválida."}), 400


# ── Histórico de Vistorias (Supabase) ─────────────────────────────────────────

@app.route("/historico/vistorias")
def historico_vistorias():
    sb = _supabase()
    vistorias = []
    erro = None
    if sb:
        try:
            res = sb.table("vistorias").select(
                "id, cliente, placa, veiculo, preenchido_por, data_hora, criado_em, arquivo_path, status, contrato_id"
            ).neq("deletado", True).order("criado_em", desc=True).execute()
            vistorias = res.data or []
        except Exception as e:
            erro = str(e)
    else:
        erro = "Supabase não configurado (SUPABASE_URL / SUPABASE_KEY ausentes)."
    return render_template("historico_vistorias.html", vistorias=vistorias, erro=erro, active="hist_vistorias")


@app.route("/historico/vistorias/<vistoria_id>/excluir", methods=["POST"])
def excluir_vistoria(vistoria_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("vistorias").update({"deletado": True}).eq("id", vistoria_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


def _reconstruir_dados_vistoria(registro: dict, sb) -> tuple:
    """
    Monta o dict `dados` para gerar_vistoria_entrada_saida a partir de um
    registro do banco. Baixa as fotos individuais do Storage quando disponíveis.
    Retorna (dados, lista_de_caminhos_temp_para_limpar).
    """
    r = registro
    temps: list[str] = []

    def _baixar_fotos(fotos_db):
        # fotos_db pode ser ["angulo:storage_path", ...] (novo) ou None/[] (sem fotos)
        result: dict[str, str] = {}
        if not sb:
            return result
        for item in (fotos_db or []):
            try:
                angulo, s_path = item.split(":", 1)
                data = sb.storage.from_("documentos").download(s_path)
                if data:
                    ext = Path(s_path).suffix.lower() or ".jpg"
                    tmp = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
                    tmp.write_bytes(bytes(data))
                    result[angulo] = str(tmp)
                    temps.append(str(tmp))
            except Exception:
                import traceback as _trc; _trc.print_exc()
        return result

    dados = {
        "contrato_id":       r.get("contrato_id", ""),
        "cliente_nome":      r.get("cliente", ""),
        "cliente_telefone":  r.get("telefone", ""),
        "cliente_endereco":  r.get("endereco", ""),
        "preenchido_por":    r.get("preenchido_por", ""),
        "veiculo":           r.get("veiculo", ""),
        "placa":             r.get("placa", ""),
        "cor":               r.get("cor", ""),
        "ano":               str(r.get("ano") or ""),
        "chassi":            r.get("chassi", ""),
        "numero_motor":      r.get("numero_motor", ""),
        # entrada — novos campos com fallback para os antigos
        "data_entrada":        r.get("data_entrada") or r.get("data_hora", ""),
        "hodometro_entrada":   r.get("hodometro_entrada") or r.get("hodometro_entrega", ""),
        "combustivel_entrada": r.get("combustivel_entrada") or r.get("combustivel", ""),
        "obs_entrada":         r.get("obs_entrada") or r.get("obs_gerais", ""),
        "sintomas_entrada":    r.get("sintomas_entrada") or r.get("desc_sintomas", ""),
        "responsavel_entrada": r.get("responsavel_entrada", ""),
        "acessorios_entrada":  r.get("acessorios_entrada") or r.get("acessorios") or {},
        "fotos_entrada":       _baixar_fotos(r.get("fotos_entrada")),
        # saída
        "data_saida":          r.get("data_saida", ""),
        "hodometro_saida":     r.get("hodometro_saida") or r.get("hodometro_retorno", ""),
        "combustivel_saida":   r.get("combustivel_saida", ""),
        "obs_saida":           r.get("obs_saida", ""),
        "sintomas_saida":      r.get("sintomas_saida", ""),
        "responsavel_saida":   r.get("responsavel_saida", ""),
        "acessorios_saida":    r.get("acessorios_saida") or {},
        "fotos_saida":         _baixar_fotos(r.get("fotos_saida")),
    }
    return dados, temps


def _gerar_docx_vistoria_bytes(registro: dict, sb) -> tuple:
    """
    Reconstrói e regenera o DOCX de uma vistoria usando o template atual.
    Retorna (docx_bytes, nome_sugerido_do_arquivo).
    """
    dados, temps = _reconstruir_dados_vistoria(registro, sb)
    placa_slug = _slugify(dados.get("placa") or "PLACA")
    nome_docx  = f"VISTORIA_{placa_slug}.docx"
    tmp_docx   = TEMP_FOLDER / f"{uuid.uuid4().hex}.docx"
    TEMP_FOLDER.mkdir(exist_ok=True)
    try:
        gerar_vistoria_entrada_saida(
            dados,
            caminho_saida=str(tmp_docx),
            template_path=str(VISTORIA_ES_TEMPLATE),
        )
        return tmp_docx.read_bytes(), nome_docx
    finally:
        tmp_docx.unlink(missing_ok=True)
        for p in temps:
            Path(p).unlink(missing_ok=True)


@app.route("/historico/vistorias/download/<vistoria_id>")
def download_vistoria_supabase(vistoria_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("vistorias").select("*").eq("id", vistoria_id).single().execute()
        registro = res.data
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    try:
        docx_bytes, nome_docx = _gerar_docx_vistoria_bytes(registro, sb)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao regenerar vistoria: {e}"}), 500
    return send_file(
        BytesIO(docx_bytes),
        as_attachment=True,
        download_name=nome_docx,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route("/historico/vistorias/<vistoria_id>/editar")
def editar_vistoria(vistoria_id):
    sb = _supabase()
    if not sb:
        flash("Supabase não configurado.", "erro")
        return redirect(url_for("historico_vistorias"))
    try:
        res = sb.table("vistorias").select("*").eq("id", vistoria_id).single().execute()
        vistoria = res.data
    except Exception as e:
        flash(f"Erro ao buscar vistoria: {e}", "erro")
        return redirect(url_for("historico_vistorias"))
    return render_template("vistoria.html", active="vistoria", vistoria=vistoria,
                           edit_id=vistoria_id, acessorios=vistoria.get("acessorios") or {})


@app.route("/historico/vistorias/download/<vistoria_id>/pdf")
def download_vistoria_pdf(vistoria_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("vistorias").select("*").eq("id", vistoria_id).single().execute()
        registro = res.data
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    try:
        docx_bytes, nome_docx = _gerar_docx_vistoria_bytes(registro, sb)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao regenerar vistoria: {e}"}), 500

    TEMP_FOLDER.mkdir(exist_ok=True)
    tmp_docx = TEMP_FOLDER / f"{uuid.uuid4().hex}.docx"
    tmp_pdf  = tmp_docx.with_suffix(".pdf")
    try:
        tmp_docx.write_bytes(docx_bytes)
        _converter_pdf(str(tmp_docx), str(tmp_pdf))
        pdf_bytes = tmp_pdf.read_bytes()
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar PDF: {e}"}), 422
    finally:
        tmp_docx.unlink(missing_ok=True)
        tmp_pdf.unlink(missing_ok=True)

    nome_pdf = nome_docx.replace(".docx", ".pdf")
    return send_file(BytesIO(pdf_bytes), as_attachment=True, download_name=nome_pdf, mimetype="application/pdf")


@app.route("/vistoria/download/<nome>")
def baixar_vistoria(nome):
    caminho = CONTRATOS_FOLDER / nome
    if not caminho.exists() or caminho.parent.resolve() != CONTRATOS_FOLDER.resolve():
        abort(404)
    ext = caminho.suffix.lower()
    mime = "application/pdf" if ext == ".pdf" else \
           "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return send_file(str(caminho), as_attachment=True, download_name=nome, mimetype=mime)


@app.route("/debug/libreoffice")
def debug_libreoffice():
    """Diagnóstico: testa conversão de um DOCX mínimo a PDF."""
    import tempfile, os, shutil, platform as _pf
    lines = [f"platform={_pf.system()}"]
    # versão do LO
    try:
        r = subprocess.run(["libreoffice", "--version"], capture_output=True, timeout=10)
        lines.append(f"version={r.stdout.decode(errors='replace').strip()}")
    except Exception as e:
        lines.append(f"version_error={e}")
    # criação de DOCX mínimo
    try:
        from docx import Document
        with tempfile.TemporaryDirectory() as wd:
            p_docx = Path(wd) / "test.docx"
            p_pdf  = Path(wd) / "test.pdf"
            doc = Document(); doc.add_paragraph("Teste LibreOffice"); doc.save(str(p_docx))
            _converter_pdf(str(p_docx), str(p_pdf))
            lines.append(f"pdf_ok={p_pdf.exists()} size={p_pdf.stat().st_size if p_pdf.exists() else 0}")
    except Exception as e:
        lines.append(f"convert_error={e}")
    return "<br>".join(lines)


# ── Controle de Inadimplência ─────────────────────────────────────────────────

def _brl(v):
    """Format float as Brazilian currency (R$ 1.234,56)."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        v = 0.0
    parts = f"{v:,.2f}".split(".")
    return "R$ " + parts[0].replace(",", ".") + "," + parts[1]


def _parse_valor_excel(raw):
    if isinstance(raw, (int, float)):
        return float(raw)
    if not raw:
        return 0.0
    s = str(raw).replace("R$", "").replace(" ", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _contratos_vencendo(dias_limite: int = 60, contratos=None):
    """Return EM ANDAMENTO contracts expiring within `dias_limite` days or already expired, sorted asc."""
    try:
        if contratos is None:
            contratos, _ = _ler_contratos()
    except Exception:
        return []
    resultado = []
    for c in contratos:
        if c.get("situacao") != "EM ANDAMENTO":
            continue
        dias_rest = c.get("dias_vencer")
        if dias_rest is None or dias_rest > dias_limite:
            continue
        resultado.append({
            "placa":          c["placa"],
            "cliente":        c["cliente"],
            "modelo":         c["modelo"],
            "termino":        c["termino_previsto"],
            "dias_restantes": dias_rest,
            "vencido":        dias_rest < 0,
        })
    resultado.sort(key=lambda x: x["dias_restantes"])
    return resultado


def _ler_inad_dados():
    """
    Fonte única: lê CONTAS-A-RECEBER.xlsx e retorna
    (registros_vencidos, registros_a_vencer, erro_leitura).
    Usada tanto por _inad_summary() quanto por pagina_inadimplencia().
    """
    from urllib.parse import quote as _url_quote
    from collections import Counter
    import openpyxl

    _base         = Path(__file__).parent / "planilhas"
    xlsx_path     = _base / "CONTAS-A-RECEBER.xlsx"
    clientes_path = _base / "DADOS_CLIENTES_CONS.xlsx"

    _tel_map = {}
    if clientes_path.exists():
        try:
            _wb_c = openpyxl.load_workbook(str(clientes_path), read_only=True, data_only=True)
            for row in _wb_c.active.iter_rows(min_row=2, values_only=True):
                nome_c = str(row[0] or "").strip()
                fone_c = str(row[11] or "").strip()
                if nome_c and fone_c:
                    digits = "".join(c for c in fone_c if c.isdigit())
                    if len(digits) >= 10:
                        _tel_map[_nh(nome_c)] = "55" + digits
            _wb_c.close()
        except Exception:
            pass

    hoje = date.today()
    registros_vencidos = []
    registros_a_vencer = []
    erro_leitura = None

    if not xlsx_path.exists():
        return registros_vencidos, registros_a_vencer, (
            "Planilha não encontrada em planilhas/. "
            "Salve o arquivo como CONTAS-A-RECEBER.xlsx nessa pasta."
        )

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header_idx = 0
        for ri, row in enumerate(rows[:10]):
            nh_row = [_nh(str(c or "")) for c in row]
            if sum(1 for t in ["receber de", "vencimento", "valor"]
                   if any(t in n for n in nh_row)) >= 2:
                header_idx = ri
                break

        header    = rows[header_idx]
        data_rows = rows[header_idx + 1:]

        def _ci(keyword):
            nk = _nh(keyword)
            return next((i for i, h in enumerate(header)
                         if h is not None and nk in _nh(str(h))), None)

        i_nome  = _ci("receber de (fantasia)") or _ci("receber de")
        i_valor = _ci("valor previsto") or _ci("valor")
        i_venc  = _ci("data de vencimento") or _ci("vencimento")
        i_sit   = _ci("situacao (data de vencimento)") or _ci("situacao")
        i_tipo  = _ci("tipo de fatura") or _ci("tipo")
        i_doc   = _ci("numero do documento") or _ci("documento")
        i_unid  = _ci("unidade")

        _NOMES_EXCLUIDOS = {"MARCELO BENTO DE ARAUJO"}

        name_counts = Counter()
        for row in data_rows:
            if i_nome is not None and i_nome < len(row) and row[i_nome]:
                n = str(row[i_nome]).strip()
                if n and n.upper() not in _NOMES_EXCLUIDOS:
                    name_counts[n] += 1

        def _get(row, idx):
            return row[idx] if idx is not None and idx < len(row) else None

        for row in data_rows:
            nome_raw = _get(row, i_nome)
            if not nome_raw:
                continue
            nome = str(nome_raw).strip()
            if not nome or nome.upper() in _NOMES_EXCLUIDOS:
                continue

            valor    = _parse_valor_excel(_get(row, i_valor))
            venc_raw = _get(row, i_venc)
            sit_raw  = _get(row, i_sit)
            tipo_raw = _get(row, i_tipo)
            doc_raw  = _get(row, i_doc)
            unid_raw = _get(row, i_unid)

            venc_date = None
            if venc_raw:
                if isinstance(venc_raw, datetime):
                    venc_date = venc_raw.date()
                elif isinstance(venc_raw, date):
                    venc_date = venc_raw
                else:
                    venc_str = str(venc_raw).strip().upper()
                    if venc_str == "HOJE":
                        venc_date = hoje
                    else:
                        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                            try:
                                venc_date = datetime.strptime(venc_str, fmt).date()
                                break
                            except (ValueError, TypeError):
                                pass
            if venc_date is None:
                continue

            situacao    = _nh(str(sit_raw  or ""))
            tipo_fatura = str(tipo_raw or "").strip()
            num_doc     = str(doc_raw  or "").strip()
            unidade     = str(unid_raw or "").strip()
            for s in ("None", "nan", ""):
                if num_doc == s: num_doc = ""
                if unidade == s: unidade = ""

            reincidente = name_counts[nome] > 1
            is_fatura   = _nh(tipo_fatura) == "fatura"
            data_fmt    = venc_date.strftime("%d/%m/%Y")

            # ── A VENCER ──────────────────────────────────────────────────────
            if "a vencer" in situacao and venc_date > hoje:
                dias_ate = (venc_date - hoje).days
                registros_a_vencer.append({
                    "nome":            nome,
                    "num_doc":         num_doc,
                    "unidade":         unidade,
                    "data_vencimento": data_fmt,
                    "dias_ate":        dias_ate,
                    "reincidente":     reincidente,
                    "tipo_fatura":     tipo_fatura,
                    "valor_s":         _brl(valor),
                    "_valor":          valor,
                })
                continue

            # ── VENCIDO ou vence hoje ─────────────────────────────────────────
            if "a vencer" in situacao and venc_date == hoje:
                dias = 0
            else:
                dias = (hoje - venc_date).days
                if dias < 0:
                    continue

            if dias == 0:    etapa, etapa_cls = "Hoje",         "stage-d0"
            elif dias == 1:  etapa, etapa_cls = "Terça-feira",  "stage-d1"
            elif dias == 2:  etapa, etapa_cls = "Quarta-feira", "stage-d2"
            elif dias == 3:  etapa, etapa_cls = "Quinta-feira", "stage-d3"
            elif dias == 4:  etapa, etapa_cls = "Sexta-feira",  "stage-d4"
            elif dias <= 6:  etapa, etapa_cls = "D+5",          "stage-d5"
            elif dias <= 9:  etapa, etapa_cls = "D+7",          "stage-d7"
            elif dias <= 14: etapa, etapa_cls = "D+10",         "stage-d10"
            else:            etapa, etapa_cls = "D+15",         "stage-d15"

            if dias == 0:    proxima = "Enviar lembrete de vencimento"
            elif dias == 1:  proxima = "Aviso de atraso — tem até o final do dia para pagar, caso contrário amanhã entram os juros"
            elif dias == 2:  proxima = "Juros aplicado — a partir de amanhã inicia a contagem dos juros de mora"
            elif dias == 3:  proxima = "Juros de mora em contagem — regularize hoje para evitar suspensão do serviço"
            elif dias == 4:  proxima = "Aviso final — regularize até hoje ou o serviço será suspenso"
            elif dias <= 6:  proxima = "Serviço suspenso — exigir comprovante de pagamento para reativação"
            elif dias <= 9:  proxima = "Encaminhar para cobrança jurídica extrajudicial"
            elif dias <= 14: proxima = "Negativação no SPC/Serasa + encaminhamento jurídico"
            else:            proxima = "Processo judicial iniciado — recolhimento imediato do veículo"

            multa      = valor * 0.10 if dias >= 2 else 0.0
            juros_mora = valor * 0.005 * dias if dias >= 3 else 0.0
            juros      = multa + juros_mora
            total      = valor + juros
            pausar     = total * 0.5

            dias_label = (f"{dias} dia{'s' if dias != 1 else ''} de atraso"
                          if dias > 0 else "Vence hoje")
            dias_s = f"{dias} dia{'s' if dias != 1 else ''}"

            if dias == 0:
                msg = f"Oi, {nome}! 😊 Passando para avisar que sua parcela de *{_brl(valor)}* vence *hoje*. Qualquer dúvida, é só chamar!\n\n\n*Ativuz Veículos*"
            elif dias == 1:
                msg = f"{nome}, sua parcela de *{_brl(valor)}* venceu ontem (vencimento: {data_fmt}). O valor atualizado é *{_brl(total)}*. Assim que puder, regularize para evitar encargos adicionais.\n\n\n*Ativuz Veículos*"
            elif dias == 2:
                msg = f"{nome}, seu pagamento de *{_brl(total)}* ainda está em aberto (vencimento: {data_fmt}). Caso tenha alguma dúvida ou dificuldade, entre em contato antes que os encargos aumentem.\n\n\n*Ativuz Veículos*"
            elif dias == 3:
                msg = f"{nome}, seu pagamento está em aberto há *{dias_s}* (vencimento: {data_fmt}). Valor atualizado: *{_brl(total)}*. Caso haja algum imprevisto, entre em contato — mas precisamos regularizar em breve para evitar a suspensão do serviço.\n\n\n*Ativuz Veículos*"
            elif dias == 4:
                msg = f"{nome}, sua parcela de *{_brl(valor)}* segue em aberto há *{dias_s}* (vencimento: {data_fmt}). Valor atualizado: *{_brl(total)}*. Regularize o quanto antes para evitar a suspensão do veículo.\n\n\n*Ativuz Veículos*"
            elif dias <= 6:
                msg = f"{nome}, infelizmente precisamos suspender o serviço por inadimplência, conforme contrato. Valor atualizado: *{_brl(total)}*. Para reativação, basta regularizar o pagamento. Estamos à disposição.\n\n\n*Ativuz Veículos*"
            elif dias <= 9:
                msg = f"{nome}, seu débito de *{_brl(total)}* está em aberto há {dias_s}. Esta é uma notificação formal com prazo de *48 horas* para regularização antes de tomarmos as próximas medidas previstas em contrato.\n\n\n*Ativuz Veículos*"
            elif dias <= 14:
                msg = f"{nome}, informamos que seu débito foi encaminhado para negativação e assessoria jurídica. Valor atualizado: *{_brl(total)}*.\n\n\n*Ativuz Veículos*"
            else:
                msg = f"{nome}, comunicamos que serão iniciados os procedimentos de protesto em cartório e execução contratual. Valor atualizado: *{_brl(total)}*.\n\n\n*Ativuz Veículos*"

            mostrar_pausar = is_fatura and 1 <= dias <= 2
            if mostrar_pausar:
                if dias == 1:
                    msg_pausar = f"{nome}, sua parcela de *{_brl(valor)}* está em aberto (vencimento: {data_fmt}). O valor atualizado é *{_brl(total)}*. 📌 Pague *{_brl(pausar)}* hoje e quite *{_brl(pausar)}* até a sexta-feira desta semana. ⚠️ Juros de 0,5% ao dia continuam correndo sobre o saldo restante. Sem pagamento até sexta, a cobrança retoma no sábado. ⚠️ Não se trata de desconto. O valor total do débito permanece integral.\n\n\n*Ativuz Veículos*"
                else:
                    msg_pausar = f"{nome}, sua parcela de *{_brl(valor)}* está em aberto há *2 dias* (vencimento: {data_fmt}). O valor atualizado é *{_brl(total)}*. 📌 Pague *{_brl(pausar)}* hoje e quite *{_brl(pausar)}* até a sexta-feira desta semana. ⚠️ Juros de 0,5% ao dia continuam correndo sobre o saldo restante. Sem pagamento até sexta, a cobrança retoma no sábado. ⚠️ Não se trata de desconto. O valor total do débito permanece integral.\n\n\n*Ativuz Veículos*"
            else:
                msg_pausar = None

            _fone = _tel_map.get(_nh(nome), "")
            wa_cobranca = f"https://wa.me/{_fone}?text=" + _url_quote(msg)
            wa_pausar   = ((f"https://wa.me/{_fone}?text=" + _url_quote(msg_pausar))
                           if mostrar_pausar else None)

            registros_vencidos.append({
                "nome":             nome,
                "num_doc":          num_doc,
                "unidade":          unidade,
                "tipo_fatura":      tipo_fatura,
                "data_vencimento":  data_fmt,
                "dias_atraso":      dias,
                "dias_label":       dias_label,
                "reincidente":      reincidente,
                "is_fatura":        is_fatura,
                "tem_multa":        dias >= 2,
                "etapa":            etapa,
                "etapa_cls":        etapa_cls,
                "proxima_acao":     proxima,
                "situacao_key":     "vence-hoje" if dias == 0 else "vencido",
                "wa_cobranca":      wa_cobranca,
                "wa_pausar":        wa_pausar,
                "msg_cobranca_txt": msg,
                "msg_pausar_txt":   msg_pausar,
                "valor_s":          _brl(valor),
                "multa_s":          _brl(multa),
                "juros_mora_s":     _brl(juros_mora),
                "juros_s":          _brl(juros),
                "total_s":          _brl(total),
                "pausar_s":         _brl(pausar),
                "_valor":           valor,
                "_multa":           multa,
                "_juros_mora":      juros_mora,
                "_juros":           juros,
                "_total":           total,
                "_fone":            _fone,
            })

    except Exception:
        import traceback; traceback.print_exc()
        erro_leitura = "Erro ao ler a planilha."

    return registros_vencidos, registros_a_vencer, erro_leitura


def _inad_summary():
    """Resumo de inadimplência para o dashboard (delega a _ler_inad_dados)."""
    try:
        registros_vencidos, _, _ = _ler_inad_dados()
    except Exception:
        return None

    if not registros_vencidos:
        return {"total_s": _brl(0), "casos": 0, "hoje": 0,
                "por_etapa": {}, "recentes": [], "total_raw": 0}

    total_raw    = sum(r["_total"] for r in registros_vencidos)
    nomes_unicos = {r["nome"] for r in registros_vencidos}
    hoje_count   = sum(1 for r in registros_vencidos if r["dias_atraso"] == 0)

    etapas = ["Hoje", "Terça-feira", "Quarta-feira", "Quinta-feira",
              "Sexta-feira", "D+5", "D+7", "D+10", "D+15"]
    por_etapa = {e: 0 for e in etapas}
    for r in registros_vencidos:
        if r["etapa"] in por_etapa:
            por_etapa[r["etapa"]] += 1

    recentes = sorted(registros_vencidos, key=lambda r: r["dias_atraso"], reverse=True)[:8]
    recentes_slim = [
        {
            "nome":    r["nome"],
            "placa":   r["num_doc"] or r["unidade"] or "—",
            "venc":    r["data_vencimento"],
            "dias":    r["dias_atraso"],
            "total_s": r["total_s"],
            "etapa":   r["etapa"],
        }
        for r in recentes
    ]

    return {
        "total_s":   _brl(total_raw),
        "total_raw": total_raw,
        "casos":     len(nomes_unicos),
        "hoje":      hoje_count,
        "por_etapa": por_etapa,
        "recentes":  recentes_slim,
    }


@app.route("/inadimplencia")
def pagina_inadimplencia():
    from collections import Counter

    registros_vencidos, registros_a_vencer, erro_leitura = _ler_inad_dados()

    registros_vencidos.sort(key=lambda r: r["dias_atraso"], reverse=True)
    registros_a_vencer.sort(key=lambda r: r["dias_ate"])

    hoje = date.today()
    total_vencidos        = len(registros_vencidos)
    total_a_vencer_cnt    = len(registros_a_vencer)
    total_valor_orig      = _brl(sum(r["_valor"] for r in registros_vencidos))
    total_valor_atual     = _brl(sum(r["_total"] for r in registros_vencidos))
    total_a_vencer_val    = _brl(sum(r["_valor"] for r in registros_a_vencer))
    criticos              = sum(1 for r in registros_vencidos if r["dias_atraso"] >= 7)
    reincidentes_criticos = sum(1 for r in registros_vencidos
                                if r["dias_atraso"] >= 7 and r["reincidente"])

    _EXCLUIR_OCORR = {"segcomp", "onevo", "new charger", "m&s", "marcelo bento de araujo"}
    _nome_cnt = Counter(
        r["nome"] for r in registros_vencidos
        if _nh(r["nome"]) not in _EXCLUIR_OCORR
    )
    _nome_valor = {}
    for r in registros_vencidos:
        if _nh(r["nome"]) not in _EXCLUIR_OCORR:
            _nome_valor[r["nome"]] = _nome_valor.get(r["nome"], 0.0) + r["_total"]

    if _nome_cnt:
        critico_ocorr_nome = _nome_cnt.most_common(1)[0][0]
        critico_ocorr_qtd  = _nome_cnt.most_common(1)[0][1]
    else:
        critico_ocorr_nome, critico_ocorr_qtd = "—", 0

    if _nome_valor:
        _cv = max(_nome_valor, key=_nome_valor.get)
        critico_valor_nome  = _cv
        critico_valor_total = _brl(_nome_valor[_cv])
    else:
        critico_valor_nome, critico_valor_total = "—", "—"

    obs_map = {}
    try:
        sb_obs = _supabase()
        if sb_obs:
            obs_res = sb_obs.table("inad_observacoes").select("chave,texto").execute()
            obs_map = {r["chave"]: r["texto"] for r in (obs_res.data or [])}
    except Exception:
        pass

    return render_template(
        "inadimplencia.html",
        registros=registros_vencidos,
        registros_a_vencer=registros_a_vencer,
        total_registros=total_vencidos,
        total_a_vencer=total_a_vencer_cnt,
        total_valor_orig=total_valor_orig,
        total_valor_atual=total_valor_atual,
        total_a_vencer_val=total_a_vencer_val,
        criticos=criticos,
        reincidentes=reincidentes_criticos,
        critico_ocorr_nome=critico_ocorr_nome,
        critico_ocorr_qtd=critico_ocorr_qtd,
        critico_valor_nome=critico_valor_nome,
        critico_valor_total=critico_valor_total,
        erro_leitura=erro_leitura,
        hoje=hoje.strftime("%d/%m/%Y"),
        active="inadimplencia",
        obs_map=obs_map,
    )


@app.route("/inadimplencia/exportar")
def exportar_inadimplencia():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border
    from openpyxl.cell.cell import MergedCell
    from collections import defaultdict

    _base      = Path(__file__).parent / "planilhas"
    xlsx_path  = _base / "CONTAS-A-RECEBER.xlsx"
    modelo     = _base / "Template_Inadimplencia.xlsx"
    hoje       = date.today()
    hoje_str   = hoje.strftime("%d/%m/%Y")
    hoje_fname = hoje.strftime("%d.%m.%y")

    # ── Re-lê e calcula registros (mesma lógica de pagina_inadimplencia) ──────
    registros = []
    if xlsx_path.exists():
        try:
            wb_src = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws_src = wb_src.active
            rows   = list(ws_src.iter_rows(values_only=True))
            wb_src.close()

            header_idx = 0
            for ri, row in enumerate(rows[:10]):
                nh_row = [_nh(str(c or "")) for c in row]
                if sum(1 for t in ["receber de","vencimento","valor"] if any(t in n for n in nh_row)) >= 2:
                    header_idx = ri; break

            header    = rows[header_idx]
            data_rows = rows[header_idx + 1:]

            def _ci(kw):
                nk = _nh(kw)
                return next((i for i,h in enumerate(header) if h and nk in _nh(str(h))), None)

            i_nome  = _ci("receber de (fantasia)") or _ci("receber de")
            i_valor = _ci("valor previsto") or _ci("valor")
            i_venc  = _ci("data de vencimento") or _ci("vencimento")
            i_sit   = _ci("situacao (data de vencimento)") or _ci("situacao")
            i_doc   = _ci("numero do documento") or _ci("documento")
            i_unid  = _ci("unidade")

            def _gv(row, idx):
                return row[idx] if idx is not None and idx < len(row) else None

            for row in data_rows:
                nome_raw = _gv(row, i_nome)
                if not nome_raw or not str(nome_raw).strip():
                    continue
                nome  = str(nome_raw).strip()
                if nome.upper() in {"MARCELO BENTO DE ARAUJO"}:
                    continue
                valor = _parse_valor_excel(_gv(row, i_valor))
                if valor <= 0:
                    continue

                venc_raw  = _gv(row, i_venc)
                sit_raw   = _nh(str(_gv(row, i_sit) or ""))
                venc_date = None
                if venc_raw:
                    if isinstance(venc_raw, datetime): venc_date = venc_raw.date()
                    elif isinstance(venc_raw, date):   venc_date = venc_raw
                    else:
                        for fmt in ["%d/%m/%Y","%Y-%m-%d","%d-%m-%Y"]:
                            try: venc_date = datetime.strptime(str(venc_raw).strip(), fmt).date(); break
                            except (ValueError, TypeError): pass
                if venc_date is None:
                    continue
                if "a vencer" in sit_raw and venc_date > hoje:
                    continue
                dias = 0 if venc_date == hoje else (hoje - venc_date).days
                if dias < 0:
                    continue

                multa      = valor * 0.10 if dias >= 2 else 0.0
                juros_mora = valor * 0.005 * dias if dias >= 3 else 0.0
                juros      = multa + juros_mora
                total      = valor + juros

                _ETAPA_SHORT = {
                    "Hoje": "Hoje", "Terça-feira": "Terça", "Quarta-feira": "Quarta",
                    "Quinta-feira": "Quinta", "Sexta-feira": "Sexta",
                    "D+5": "D+5", "D+7": "D+7", "D+10": "D+10", "D+15": "D+15",
                }
                if   dias == 0:     etapa, proxima = "Hoje",         "Enviar lembrete de vencimento"
                elif dias == 1:     etapa, proxima = "Terça-feira",  "Aviso de atraso — tem até o final do dia para pagar, caso contrário amanhã entram os juros"
                elif dias == 2:     etapa, proxima = "Quarta-feira", "Juros aplicado — a partir de amanhã inicia a contagem dos juros de mora"
                elif dias == 3:     etapa, proxima = "Quinta-feira", "Juros de mora em contagem — regularize hoje para evitar suspensão do serviço"
                elif dias == 4:     etapa, proxima = "Sexta-feira",  "Aviso final — regularize até hoje ou o serviço será suspenso"
                elif dias <= 6:     etapa, proxima = "D+5",  "Serviço suspenso — exigir comprovante de pagamento para reativação"
                elif dias <= 9:     etapa, proxima = "D+7",  "Encaminhar para cobrança jurídica extrajudicial"
                elif dias <= 14:    etapa, proxima = "D+10", "Negativação no SPC/Serasa + encaminhamento jurídico"
                else:               etapa, proxima = "D+15", "Processo judicial iniciado — recolhimento imediato do veículo"

                registros.append({
                    "nome": nome, "etapa": etapa, "etapa_short": _ETAPA_SHORT.get(etapa, etapa),
                    "proxima": proxima, "vencimento": venc_date.strftime("%d/%m/%Y"), "dias": dias,
                    "valor": valor, "juros": juros, "total": total,
                })
        except Exception:
            import traceback; traceback.print_exc()

    # ── Gera Excel ────────────────────────────────────────────────────────────
    def _fill(hex6): return PatternFill("solid", fgColor=hex6)
    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    F_ROW_ODD = _fill("F0F4FA")
    F_ROW_EVN = _fill("FFFFFF")
    F_VALOR   = _fill("EFF6FF")
    F_JUROS   = _fill("FFF7ED")
    F_TOTAL_C = _fill("1E3A5F")
    FMT_BRL   = '"R$"\\ #,##0.00'

    # Cor de fundo por etapa (coluna Etapa)
    _ETAPA_BG = {
        "Hoje": "2563EB", "Terça": "1D4ED8", "Quarta": "1E40AF",
        "Quinta": "D97706", "Sexta": "B45309",
        "D+5": "DC2626", "D+7": "B91C1C", "D+10": "991B1B", "D+15": "7F1D1D",
    }
    _ETAPA_PRIO = {
        "Hoje": 1, "Terça": 2, "Quarta": 3, "Quinta": 4, "Sexta": 5,
        "D+5": 6, "D+7": 7, "D+10": 8, "D+15": 9,
    }

    F_NONE = PatternFill(fill_type=None)  # sem preenchimento (limpa fill)

    def _safe_set(cell, **kwargs):
        if isinstance(cell, MergedCell):
            return
        for attr, val in kwargs.items():
            setattr(cell, attr, val)

    def _unmerge_area(ws, min_row, max_row, min_col, max_col):
        to_remove = [
            m for m in list(ws.merged_cells.ranges)
            if m.min_row <= max_row and m.max_row >= min_row
            and m.min_col <= max_col and m.max_col >= min_col
        ]
        for m in to_remove:
            ws.unmerge_cells(str(m))

    def _clear_rows(ws, from_row, to_row, min_col, max_col):
        """Limpa valor e fill de células numa faixa de linhas."""
        _unmerge_area(ws, from_row, to_row, min_col, max_col)
        for r in range(from_row, to_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill  = F_NONE

    wb = openpyxl.load_workbook(str(modelo))

    total_valor = sum(r["valor"] for r in registros)
    total_juros = sum(r["juros"] for r in registros)
    total_total = sum(r["total"] for r in registros)
    n = len(registros)

    F_TOT_FONT = Font(color="FFFFFF", bold=True, size=10)

    # ── Aba 1: Resumo Executivo ───────────────────────────────────────────────
    ws1 = wb["Resumo Executivo"]

    _safe_set(ws1["B3"],
              value=f"ATIVUZ VEÍCULOS  ·  Gerado em: {hoje_str}  ·  Todos os títulos — ordem por maior valor")

    D_INI = 10
    D_LIM = 300
    registros_val = sorted(registros, key=lambda x: x["valor"], reverse=True)
    D_FIM = D_INI + n - 1 if n > 0 else D_INI

    _clear_rows(ws1, D_INI, D_LIM, 2, 8)

    for i, rec in enumerate(registros_val):
        r    = D_INI + i
        base = F_ROW_ODD if i % 2 == 0 else F_ROW_EVN
        es   = rec["etapa_short"]
        F_ET = _fill(_ETAPA_BG.get(es, "374151"))
        F_EF = Font(color="FFFFFF", bold=True, size=9)

        _safe_set(ws1.cell(r, 2), value=rec["nome"],       fill=base,    font=Font(size=10),              alignment=_align("left"))
        _safe_set(ws1.cell(r, 3), value=es,                fill=F_ET,    font=F_EF,                       alignment=_align("center"))
        _safe_set(ws1.cell(r, 4), value=rec["vencimento"], fill=base,    font=Font(size=10),              alignment=_align("center"))
        _safe_set(ws1.cell(r, 5), value=rec["dias"],       fill=base,    font=Font(size=10),              alignment=_align("center"))
        _safe_set(ws1.cell(r, 6), value=rec["valor"],      fill=F_VALOR, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws1.cell(r, 7), value=rec["juros"],      fill=F_JUROS, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws1.cell(r, 8), value=rec["total"],      fill=F_ROW_EVN, font=Font(bold=True, size=10), number_format=FMT_BRL, alignment=_align("right"))

    T_ROW1 = D_FIM + 1
    _safe_set(ws1.cell(T_ROW1, 2), value="TOTAL", fill=F_TOTAL_C, font=F_TOT_FONT, alignment=_align("center"))
    for col in [3, 4, 5]:
        _safe_set(ws1.cell(T_ROW1, col), fill=F_TOTAL_C)
    _safe_set(ws1.cell(T_ROW1, 6), value=f"=SUM(F{D_INI}:F{D_FIM})" if n else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws1.cell(T_ROW1, 7), value=f"=SUM(G{D_INI}:G{D_FIM})" if n else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws1.cell(T_ROW1, 8), value=f"=SUM(H{D_INI}:H{D_FIM})" if n else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))

    # Exclui linhas abaixo do TOTAL
    if T_ROW1 + 1 <= D_LIM:
        ws1.delete_rows(T_ROW1 + 1, D_LIM - T_ROW1)

    _safe_set(ws1["B7"], value=n)
    _safe_set(ws1["C7"], value=total_valor, number_format=FMT_BRL)
    _safe_set(ws1["F7"], value=total_juros, number_format=FMT_BRL)
    _safe_set(ws1["H7"], value=total_total, number_format=FMT_BRL)

    # ── Aba 2: Detalhamento por Cliente ──────────────────────────────────────
    ws2 = wb["Detalhamento por Cliente"]

    _safe_set(ws2["B3"],
              value=f"Valores consolidados por cliente — ordem alfabética  ·  {hoje_str}")

    clientes_map = defaultdict(lambda: {"etapa_pior": "", "prio": 0, "titulos": 0, "valor": 0.0, "juros": 0.0})
    for rec in registros:
        c = clientes_map[rec["nome"]]
        c["titulos"] += 1
        c["valor"]   += rec["valor"]
        c["juros"]   += rec["juros"]
        p = _ETAPA_PRIO.get(rec["etapa_short"], 0)
        if p > c["prio"]:
            c["prio"] = p
            c["etapa_pior"] = rec["etapa_short"]

    clientes_sorted = sorted(clientes_map.items(), key=lambda x: x[0])
    nc = len(clientes_sorted)

    D_INI2 = 10
    D_FIM2 = D_INI2 + nc - 1 if nc > 0 else D_INI2

    _clear_rows(ws2, D_INI2, D_LIM, 2, 7)

    for i, (nome, g) in enumerate(clientes_sorted):
        r    = D_INI2 + i
        base = F_ROW_ODD if i % 2 == 0 else F_ROW_EVN
        F_ET = _fill(_ETAPA_BG.get(g["etapa_pior"], "374151"))
        F_EF = Font(color="FFFFFF", bold=True, size=9)
        total_cli = g["valor"] + g["juros"]

        _safe_set(ws2.cell(r, 2), value=nome,            fill=base,    font=Font(size=10),              alignment=_align("left"))
        _safe_set(ws2.cell(r, 3), value=g["etapa_pior"], fill=F_ET,    font=F_EF,                       alignment=_align("center"))
        _safe_set(ws2.cell(r, 4), value=g["titulos"],    fill=base,    font=Font(size=10),              alignment=_align("center"))
        _safe_set(ws2.cell(r, 5), value=g["valor"],      fill=F_VALOR, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws2.cell(r, 6), value=g["juros"],      fill=F_JUROS, font=Font(size=10), number_format=FMT_BRL, alignment=_align("right"))
        _safe_set(ws2.cell(r, 7), value=total_cli,       fill=F_ROW_EVN, font=Font(bold=True, size=10), number_format=FMT_BRL, alignment=_align("right"))

    T_ROW2 = D_FIM2 + 1
    _safe_set(ws2.cell(T_ROW2, 2), value="TOTAL GERAL", fill=F_TOTAL_C, font=F_TOT_FONT, alignment=_align("center"))
    for col in [3, 4]:
        _safe_set(ws2.cell(T_ROW2, col), fill=F_TOTAL_C)
    _safe_set(ws2.cell(T_ROW2, 5), value=f"=SUM(E{D_INI2}:E{D_FIM2})" if nc else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws2.cell(T_ROW2, 6), value=f"=SUM(F{D_INI2}:F{D_FIM2})" if nc else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws2.cell(T_ROW2, 7), value=f"=SUM(G{D_INI2}:G{D_FIM2})" if nc else 0,
              fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))

    # Exclui linhas abaixo do TOTAL
    if T_ROW2 + 1 <= D_LIM:
        ws2.delete_rows(T_ROW2 + 1, D_LIM - T_ROW2)

    _safe_set(ws2["B7"], value=nc)
    _safe_set(ws2["D7"], value=total_valor, number_format=FMT_BRL)
    _safe_set(ws2["F7"], value=total_total, number_format=FMT_BRL)

    # ── Aba 3: Análise por Etapa ──────────────────────────────────────────────
    ws3 = wb["Análise por Etapa"]

    _safe_set(ws3["B3"],
              value=f"Resumo consolidado por etapa de cobrança  ·  {hoje_str}")

    ETAPA_ROWS3 = {
        "Hoje": 10, "Terça": 11, "Quarta": 12, "Quinta": 13, "Sexta": 14,
        "D+5": 15, "D+7": 16, "D+10": 17, "D+15": 18,
    }
    agrup = defaultdict(lambda: {"n": 0, "valor": 0.0, "juros": 0.0})
    for rec in registros:
        g = agrup[rec["etapa_short"]]
        g["n"] += 1; g["valor"] += rec["valor"]; g["juros"] += rec["juros"]

    for etapa_s, row_num in ETAPA_ROWS3.items():
        g = agrup[etapa_s]
        total_e = g["valor"] + g["juros"]
        _safe_set(ws3.cell(row_num, 3), value=g["n"],      alignment=_align("center"))
        _safe_set(ws3.cell(row_num, 4), value=g["valor"],  number_format=FMT_BRL, fill=F_VALOR)
        _safe_set(ws3.cell(row_num, 5), value=g["juros"],  number_format=FMT_BRL, fill=F_JUROS)
        _safe_set(ws3.cell(row_num, 6), value=total_e,     number_format=FMT_BRL)
        _safe_set(ws3.cell(row_num, 9), value=total_e,     number_format=FMT_BRL)

    # Total row (row 19 já tem label "TOTAL" no template)
    _safe_set(ws3.cell(19, 3), value="=SUM(C10:C18)", fill=F_TOTAL_C, font=F_TOT_FONT, alignment=_align("center"))
    _safe_set(ws3.cell(19, 4), value="=SUM(D10:D18)", fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws3.cell(19, 5), value="=SUM(E10:E18)", fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))
    _safe_set(ws3.cell(19, 6), value="=SUM(F10:F18)", fill=F_TOTAL_C, font=F_TOT_FONT, number_format=FMT_BRL, alignment=_align("right"))

    etapas_ativas = sum(1 for g in agrup.values() if g["n"] > 0)
    _safe_set(ws3["B7"], value=etapas_ativas)
    _safe_set(ws3["C7"], value=n)
    _safe_set(ws3["E7"], value=total_total, number_format=FMT_BRL)

    # ── Serve o arquivo ───────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f"Relatório_Inadimplência_{hoje_fname}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/inadimplencia/upload", methods=["POST"])
def inadimplencia_upload():
    f = request.files.get("planilha")
    if not f or not f.filename:
        flash("Nenhum arquivo selecionado.", "error")
        return redirect(url_for("pagina_inadimplencia"))
    if not f.filename.lower().endswith(".xlsx"):
        flash("Apenas arquivos .xlsx são aceitos.", "error")
        return redirect(url_for("pagina_inadimplencia"))
    dest = Path(__file__).parent / "planilhas" / "CONTAS-A-RECEBER.xlsx"
    f.save(str(dest))
    flash("Planilha atualizada! Os dados abaixo refletem o arquivo enviado.", "success")
    return redirect(url_for("pagina_inadimplencia"))


_MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
_MESES_PT_CURTO = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                   "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _nome_mes_label(mes, ano, acumulado=False):
    if acumulado:
        return f"Jan–{_MESES_PT_CURTO[mes - 1]} {ano}"
    return f"{_MESES_PT[mes - 1]} {ano}"


# ── Configurações DRE ────────────────────────────────────────────────────────
# IR/CSLL: manter 0.0 para Simples Nacional; alterar aqui para mudar regime.
_DRE_IR_CSLL = 0.0

# Códigos cujo Tipo no sistema (ENTRADA/SAÍDA) é oposto ao tratamento no DRE.
# Ao adicionar novos, revisar o grupo de destino no _DRE_LAYOUT abaixo.
_DRE_CODIGOS_SINAL_INVERTIDO = frozenset([
    "02.02.06.006",  # SAÍDA → receita operacional (Combustível Reembolsável)
    "01.01.02.008",  # ENTRADA → dedução (Desconto Concedido a Clientes)
    "02.04.06.004",  # SAÍDA → receita financeira (Desconto Pgto Boletos)
])

# ── DRE Layout ────────────────────────────────────────────────────────────────
# sign: +1 = receita, -1 = despesa.  item tuple: (codigo, label).
# Códigos tratados como string — nunca converter para número.

_DRE_LAYOUT = [
    {"id": "rb", "label": "(+) Receita Operacional Bruta", "grupos": [
        {"id": "rb-loc",  "label": "Locação", "sign": +1, "itens": [
            ("01.01.01.001", "Locação"),
            ("01.01.01.002", "KM Excedente"),
            ("01.01.01.003", "Multa Atraso Pagamento"),
            ("01.01.01.004", "Multa Quebra de Contrato"),
            ("01.01.01.005", "Acordo/Renegociação"),
            ("01.01.01.006", "Taxa de Adm. de Veículos"),
        ]},
        {"id": "rb-reimb", "label": "Reembolsos", "sign": +1, "itens": [
            ("01.01.02.001", "Manutenção Reembolsável"),
            ("01.01.02.002", "Entrada de Multa de Trânsito"),
            ("01.01.02.004", "Multa Dev. Antecipada"),
            ("01.01.02.005", "Reembolso de Sinistro"),
            ("01.01.02.006", "Outros Reembolsos"),
            ("02.02.06.006", "Combustível Reembolsável"),  # SAÍDA no sistema → sinal invertido
        ]},
        {"id": "rb-fi", "label": "Reembolsos — Frota Investidores", "sign": +1, "itens": [
            ("01.01.02.009", "Reembolso de Manutenções — FI"),
            ("01.01.02.010", "Reembolso de Multas — FI"),
            ("01.01.02.011", "Reembolso Desp. Operacionais — FI"),
            ("01.02.01.005", "Recebimentos — Frota Investidores"),
        ]},
    ]},
    {"id": "ded", "label": "(-) Deduções", "grupos": [
        {"id": "ded-imp",  "label": "Impostos", "sign": -1,
         "note": "incide sobre o faturamento — Simples Nacional", "itens": [
            ("02.01.01.001", "PIS"),
            ("02.01.01.005", "Simples Nacional"),
            ("02.01.01.006", "Outros Impostos"),
        ]},
        {"id": "ded-desc", "label": "Descontos", "sign": -1, "itens": [
            ("01.01.02.008", "Desconto Concedido a Clientes"),  # ENTRADA → sinal invertido
        ]},
    ]},
    # subtotal: receita_liquida
    {"id": "custos", "label": "(-) Custos — Custo Direto da Operação/Frota", "grupos": [
        {"id": "c-lic",   "label": "Licenciamento", "sign": -1, "itens": [
            ("02.02.01.001", "Emplacamento"),
            ("02.02.01.002", "Transferência Veicular"),
            ("02.02.01.003", "IPVA"),
            ("02.02.01.005", "Taxa de Licenciamento"),
            ("02.02.01.006", "Despesas com Cartório"),
            ("02.02.01.007", "Transferência Veicular"),
            ("02.02.01.008", "Taxa Bombeiros"),
            ("02.02.01.009", "Vistoria"),
            ("02.02.01.010", "Documentação Veicular"),
        ]},
        {"id": "c-desp",  "label": "Honorários Despachante", "sign": -1, "itens": [
            ("02.02.02.001", "Honorários Despachante"),
        ]},
        {"id": "c-seg",   "label": "Seguro/Assistência 24h", "sign": -1, "itens": [
            ("02.02.03.001", "Seguro Total"),
            ("02.02.03.003", "Reserva Operacional p/ Sinistros"),
            ("02.02.03.004", "Guincho"),
            ("02.02.03.005", "Rastreador Veicular"),
        ]},
        {"id": "c-sub",   "label": "Sublocação/Transporte", "sign": -1, "itens": [
            ("02.02.04.002", "Frete Não Reembolsável"),
            ("02.02.04.003", "Multas Não Reembolsáveis"),
            ("02.02.04.004", "Taxa de Frete"),
            ("02.02.04.005", "Gasolina"),
            ("02.02.04.006", "Uber ou App de Transporte"),
        ]},
        {"id": "c-man",   "label": "Manutenção", "sign": -1, "itens": [
            ("02.02.05.001", "Manutenção Preventiva"),
            ("02.02.05.002", "Manutenção Corretiva"),
            ("02.02.05.006", "Compra Equipamento GNV"),
            ("02.02.05.007", "Equipamentos (Sensor, SIM, etc.)"),
            ("02.02.05.008", "Lavagem Veicular Não Reembolsável"),
            ("02.02.05.009", "Compra de Peças"),
            ("02.02.05.010", "Compra de Pneus"),
        ]},
        {"id": "c-reimb", "label": "Despesas Reembolsáveis", "sign": -1, "itens": [
            ("02.02.06.001", "Saída de Multa de Trânsito"),
            ("02.02.06.002", "Saída de Sinistro"),
            ("02.02.06.010", "Reembolso de Clientes"),
        ]},
        {"id": "c-fi",    "label": "Despesas c/ Frota de Investidores", "sign": -1, "itens": [
            ("02.02.07.01", "Manutenções — Frota Investidores"),
            ("02.02.07.02", "Sinistros — Frota Investidores"),
            ("02.02.07.03", "Desp. Operacionais — FI"),
            ("02.02.07.04", "Desp. Operacionais — FI"),
        ]},
    ]},
    # subtotal: margem
    {"id": "sga", "label": "(-) SG&A — Despesas Gerais e Administrativas", "grupos": [
        {"id": "s-sal", "label": "Salários", "sign": -1, "itens": [
            ("02.03.01.001", "Salário"),
            ("02.03.01.002", "Adiantamento Salarial"),
            ("02.03.01.003", "Férias"),
            ("02.03.01.004", "13° Salário"),
            ("02.03.01.005", "Rescisão"),
            ("02.03.01.006", "Prêmios"),
            ("02.03.01.007", "Comissão"),
            ("02.03.01.10",  "ASO e Saúde e Seg. do Trabalho"),
        ]},
        {"id": "s-ben", "label": "Benefícios", "sign": -1, "itens": [
            ("02.03.02.001", "VT"),
            ("02.03.02.003", "VR"),
            ("02.03.02.005", "Assistência Médica"),
            ("02.03.02.006", "P.C.M.S.O"),
            ("02.03.02.007", "Treinamento"),
            ("02.03.02.008", "Outros Benefícios"),
        ]},
        {"id": "s-imp", "label": "Impostos Folha", "sign": -1, "itens": [
            ("02.03.03.001", "INSS"),
            ("02.03.03.002", "FGTS"),
        ]},
        {"id": "s-pro", "label": "Pró-labore", "sign": -1,
         "highlight": True, "note": "remuneração dos sócios", "itens": [
            ("02.03.04.001", "Pró-labore Folha"),
        ]},
        {"id": "s-com", "label": "Despesas Comerciais", "sign": -1, "itens": [
            ("02.04.01.001", "Marketing"),
        ]},
        {"id": "s-ocu", "label": "Ocupação", "sign": -1, "itens": [
            ("02.04.02.001", "Aluguel de Imóveis"),
            ("02.04.02.003", "IPTU"),
            ("02.04.02.004", "Água"),
            ("02.04.02.005", "Luz"),
            ("02.04.02.006", "Manutenção Predial"),
        ]},
        {"id": "s-sup", "label": "Suprimentos", "sign": -1, "itens": [
            ("02.04.03.001", "Telefone"),
            ("02.04.03.007", "Bens de Pequeno Valor"),
        ]},
        {"id": "s-adm", "label": "Despesas Administrativas", "sign": -1, "itens": [
            ("02.04.04.001", "Desp. Administrativas e de Escritório"),
            ("02.04.04.004", "Taxas e Despesas Legais"),
            ("02.04.04.005", "Outras Despesas"),
            ("02.04.04.007", "Despesas Jurídicas"),
        ]},
        {"id": "s-svc", "label": "Serviços Prestados", "sign": -1, "itens": [
            ("02.04.05.001", "Honorários Advocatícios"),
            ("02.04.05.003", "Softwares"),
            ("02.04.05.004", "Órgãos de Proteção ao Crédito"),
            ("02.04.05.005", "Assessoria Administrativa"),
            ("02.04.05.006", "Honorários de Consultoria"),
            ("02.04.05.007", "Serviços Contábeis"),
            ("02.04.05.008", "Serviços de Limpeza"),
            ("02.04.05.009", "Serviços Manut. Máquinas e Equip."),
        ]},
        {"id": "s-out", "label": "Outras Saídas", "sign": -1, "itens": [
            ("02.04.07.002", "Outras Saídas"),
        ]},
    ]},
    # subtotal: ebitda; depreciação=0; subtotal: ebit
    {"id": "rfin", "label": "(-) Resultado Financeiro", "grupos": [
        {"id": "rf-desp", "label": "Despesas Bancárias e Financeiras", "sign": -1, "itens": [
            ("02.04.06.001", "Tarifa Bancária"),
            ("02.04.06.003", "Juros e Multas Bancárias Pagos"),
            ("02.04.06.005", "Taxa Maquineta"),
            ("03.01.03.002", "Consórcio Contemplado Juros"),
            ("03.01.03.004", "Financiamento Juros"),
            ("04.01.02.002", "Pgto Juros sobre Mútuos"),
        ]},
        {"id": "rf-rec",  "label": "Receitas Financeiras", "sign": +1, "itens": [
            ("03.03.01.003", "Rendimento de Aplicações"),
            ("02.04.06.004", "Desconto Pgto Boletos"),  # SAÍDA no sistema → receita financeira
        ]},
    ]},
    {"id": "rnop", "label": "(-) Resultados Não Operacionais", "grupos": [
        {"id": "rnop-out", "label": "Outras Entradas", "sign": +1, "itens": [
            ("01.02.01.002", "Depósitos Não Identificados"),
            ("01.02.01.003", "Outras Entradas"),
        ]},
        {"id": "rnop-inv", "label": "Outros Investimentos", "sign": +1, "itens": [
            ("03.03.02.001", "Outros Investimentos"),
        ]},
    ]},
    # subtotal: lucro_liquido ── abaixo: fluxo de caixa
    {"id": "inv", "label": "(-) Investimentos", "grupos": [
        {"id": "inv-venda", "label": "Venda de Veículos", "sign": +1, "itens": [
            ("01.02.01.004", "Venda de Veículo"),
        ]},
        {"id": "inv-comp",  "label": "Compra de Veículos", "sign": -1, "itens": [
            ("03.01.02.001", "Compra de Veículos à Vista"),
            ("03.01.02.002", "Entrada Compra Veículo"),
            ("03.01.02.003", "Adiantamento de Consórcio"),
        ]},
        {"id": "inv-fin",   "label": "Financiamentos Veiculares", "sign": -1, "itens": [
            ("03.01.03.001", "Consórcio Contemplado"),
            ("03.01.03.003", "Pagamento de Financiamento"),
            ("03.01.03.005", "Consórcio Parcela Não Contemplada"),
            ("03.01.03.006", "Quitação Antecipada de Parcelas"),
        ]},
        {"id": "inv-imob",  "label": "Outros Imobilizados", "sign": -1, "itens": [
            ("03.02.01.001", "Instalações"),
            ("03.02.01.002", "Computadores e Periféricos"),
            ("03.02.01.003", "Móveis e Utensílios"),
            ("03.02.01.004", "Sistemas e Softwares"),
            ("03.02.01.005", "Outras Imobilizações"),
        ]},
        {"id": "inv-obra",  "label": "Construção da Oficina", "sign": -1, "itens": [
            ("02.04.08.01", "Compra de Material para Oficina"),
            ("02.04.08.02", "Pagamento da Mão de Obra"),
            ("02.04.08.03", "Aluguel de Equipamentos"),
        ]},
        {"id": "inv-aplic", "label": "Aplicações Financeiras", "sign": -1, "itens": [
            ("03.03.01.001", "Aplicações Financeiras"),
        ]},
        {"id": "inv-resg",  "label": "Resgate de Aplicação", "sign": +1, "itens": [
            ("03.03.01.002", "Resgate de Aplicação Financeira"),
        ]},
    ]},
    {"id": "financ", "label": "(-) Financiamentos", "grupos": [
        {"id": "fin-ent",      "label": "Entradas de Mútuos", "sign": +1, "itens": [
            ("04.01.01.001", "Entrada de Mútuos"),
            ("04.01.01.002", "Entrada Caução"),
        ]},
        {"id": "fin-pgto",     "label": "Pgto de Mútuos", "sign": -1, "itens": [
            ("04.01.02.001", "Saída de Mútuos"),
            ("04.01.02.003", "Saída Caução"),
        ]},
        {"id": "fin-reimb-in", "label": "Entrada de Reembolso", "sign": +1, "itens": [
            ("04.04.02.01", "Entrada de Reembolso"),
        ]},
        {"id": "fin-reimb-out","label": "Saída de Reembolso", "sign": -1, "itens": [
            ("04.04.02.02", "Saída de Reembolso"),
        ]},
    ]},
    {"id": "aporte", "label": "(+) Aporte de Capital", "grupos": [
        {"id": "aporte-g", "label": "Aporte de Capital", "sign": +1, "itens": [
            ("04.04.01.003", "Aporte de Capital"),
        ]},
    ]},
    # subtotal: fluxo_acionista
    {"id": "distrib", "label": "(-) Distribuição de Resultado", "grupos": [
        {"id": "distrib-g",     "label": "Distribuição de Resultado", "sign": -1, "itens": [
            ("04.04.01.002", "Distribuição de Resultado"),
        ]},
        {"id": "distrib-lucro", "label": "Distribuição de Lucros Mensal", "sign": -1, "itens": [
            ("02.03.04.002", "Distribuição de Lucros Mensal"),
        ]},
    ]},
    # subtotal: fluxo_livre
]


def _dre_ler_lancamentos():
    import openpyxl

    base = Path(__file__).resolve().parent
    pasta = base / "planilhas" / "dre"

    arquivos = sorted(pasta.glob("*.xlsx")) if pasta.is_dir() else []
    if not arquivos:
        return []

    def _split(s):
        if ' - ' in s:
            a, b = s.split(' - ', 1)
            return a.strip(), b.strip()
        return s.strip(), s.strip()

    def _ler_arquivo(path):
        registros = []
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            for row in rows[5:]:
                if not row[0]:
                    continue
                dt    = row[3]
                valor = row[9]
                if not isinstance(dt, datetime) or not isinstance(valor, (int, float)):
                    continue
                cod, _ = _split(str(row[2]) if row[2] else "")
                num = str(row[6]).strip() if row[6] else ""
                registros.append({"num": num, "codigo": cod, "dt": dt, "valor": float(valor)})
        except Exception:
            pass
        return registros

    # Combina todos os arquivos e deduplica por (número do lançamento, código de natureza)
    vistos = set()
    result = []
    for arq in sorted(arquivos, key=lambda p: p.stat().st_mtime):
        for reg in _ler_arquivo(arq):
            chave = (reg["num"], reg["codigo"])
            if chave in vistos:
                continue
            vistos.add(chave)
            result.append(reg)
    return result


def _dre_calcular(lancamentos):
    from collections import defaultdict
    code_val = defaultdict(float)
    for l in lancamentos:
        if l["codigo"]:
            code_val[l["codigo"]] += l["valor"]

    sections = []
    for sec_def in _DRE_LAYOUT:
        grupos = []
        sec_total = 0.0
        for grp_def in sec_def["grupos"]:
            sign = grp_def["sign"]
            itens = []
            grp_abs = 0.0
            for codigo, label in grp_def["itens"]:
                v = code_val.get(codigo, 0.0)
                grp_abs += v
                itens.append({"codigo": codigo, "label": label, "val": sign * v})
            grp_total = sign * grp_abs
            sec_total += grp_total
            grupos.append({
                "id": grp_def["id"], "label": grp_def["label"],
                "sign": sign, "total": grp_total, "itens": itens,
                "highlight": grp_def.get("highlight", False),
                "note": grp_def.get("note", ""),
            })
        sections.append({"id": sec_def["id"], "label": sec_def["label"],
                         "total": sec_total, "grupos": grupos})

    def _s(sid):
        for s in sections:
            if s["id"] == sid:
                return s["total"]
        return 0.0

    rb     = _s("rb")
    ded    = _s("ded")
    rl     = rb + ded
    custos = _s("custos")
    margem = rl + custos
    sga    = _s("sga")
    ebitda = margem + sga
    ebit   = ebitda        # depreciação = 0
    rfin   = _s("rfin")
    rnop   = _s("rnop")
    lucro_antes_ir = ebit + rfin + rnop
    ir_csll = _DRE_IR_CSLL
    ll     = lucro_antes_ir - ir_csll
    inv    = _s("inv")
    financ = _s("financ")
    aporte = _s("aporte")
    fluxo_ac  = ll + inv + financ + aporte
    distrib   = _s("distrib")
    fluxo_liv = fluxo_ac + distrib

    # Add %RL to each section (informational; template may display for L1 rows)
    for sec in sections:
        sec["pct"] = sec["total"] / rl if rl else 0.0

    return {
        "sections": sections,
        "receita_bruta": rb,  "deducoes": ded,
        "receita_liquida": rl,
        "custos": custos,
        "margem": margem,       "pct_margem": margem / rl if rl else 0,
        "sga": sga,
        "ebitda": ebitda,       "pct_ebitda": ebitda / rl if rl else 0,
        "depreciacao": 0.0,
        "ebit": ebit,           "pct_ebit": ebit / rl if rl else 0,
        "rfin": rfin,           "rnop": rnop,
        "lucro_antes_ir": lucro_antes_ir,
        "pct_lajir": lucro_antes_ir / rl if rl else 0,
        "ir_csll": ir_csll,
        "lucro_liquido": ll,    "pct_ll": ll / rl if rl else 0,
        "inv": inv,  "financ": financ,  "aporte": aporte,
        "fluxo_acionista": fluxo_ac,
        "distrib": distrib,
        "fluxo_livre": fluxo_liv,
    }


def _ler_lancamentos_jun_jul():
    """Lê o arquivo histórico de junho/julho 2025 (formato diferente do DRE principal)."""
    import openpyxl
    path = Path(__file__).resolve().parent / "planilhas" / "dados_junho_julho.xlsx"
    if not path.exists():
        return []
    registros = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        for row in rows[1:]:  # pula cabeçalho
            if not row[0]:
                continue
            try:
                dt = datetime.strptime(str(row[2]), "%d/%m/%Y")
                natureza = str(row[3]) if row[3] else ""
                cod = natureza.split(" - ")[0].strip() if " - " in natureza else natureza.strip()
                valor_str = str(row[4]).replace("R$", "").replace("-", "").replace(".", "").replace(",", ".").strip()
                valor = float(valor_str)
                registros.append({"codigo": cod, "dt": dt, "valor": valor})
            except Exception:
                continue
    except Exception:
        pass
    return registros


def _calcular_indicadores_ativuz():
    """Calcula margens da Ativuz usando os últimos 12 meses de lançamentos."""
    from calendar import monthrange
    hoje = datetime.now(_BRT)
    mes_ini = hoje.month + 1
    ano_ini = hoje.year - 1
    if mes_ini > 12:
        mes_ini -= 12
        ano_ini += 1
    d_ini = datetime(ano_ini, mes_ini, 1)
    d_fim = datetime(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1], 23, 59, 59)

    todos = _dre_ler_lancamentos() + _ler_lancamentos_jun_jul()
    filtrados = [l for l in todos if d_ini <= l["dt"] <= d_fim]
    if not filtrados:
        return None

    dre = _dre_calcular(filtrados)

    # Saldo devedor da carteira financeira (dívida líquida)
    saldo_devedor = 0.0
    try:
        sb = _supabase()
        if sb:
            rows = sb.table("financiamentos_contratos").select("*").execute().data or []
            hoje_d = hoje.date()
            from math import ceil
            for r in rows:
                try:
                    vcto = date.fromisoformat(str(r.get("data_vencimento", ""))[:10])
                    dias = (vcto - hoje_d).days
                    restante = ceil(dias / 30.44) if dias > 0 else 0
                    saldo_devedor += restante * float(r["valor_parcela"])
                except Exception:
                    continue
    except Exception:
        pass

    # Depreciação anual da frota (vida útil 5 anos — Receita Federal veículos leves)
    depreciacao_anual = 0.0
    try:
        sb2 = _supabase()
        if sb2:
            res_frota = sb2.table("frota_veiculos").select("vl_aquisicao").eq("ativo", True).execute()
            for v in (res_frota.data or []):
                try:
                    depreciacao_anual += float(v.get("vl_aquisicao") or 0) / 5
                except (TypeError, ValueError):
                    continue
    except Exception:
        pass

    def _pct(v):
        return f"{v * 100:.2f}%".replace(".", ",")

    def _ratio(numerador, denominador):
        if not denominador:
            return "N/D"
        return f"{numerador / denominador:.2f}".replace(".", ",")

    rl     = dre["receita_liquida"]
    ebitda = dre["ebitda"]
    ebit   = ebitda - depreciacao_anual
    pct_ebit = (ebit / rl) if rl else 0

    return {
        "ticker": "ATIVUZ", "nome": "Ativuz", "erro": None, "is_ativuz": True,
        "pl":             "N/A",
        "pvp":            "N/A",
        "roe":            "N/D",
        "margem_bruta":   _pct(dre["pct_margem"]),
        "margem_ebitda":  _pct(dre["pct_ebitda"]),
        "margem_ebit":    _pct(pct_ebit),
        "margem_liquida": _pct(dre["pct_ll"]),
        "div_ebitda":     _ratio(saldo_devedor, ebitda),
        "div_ebit":       _ratio(saldo_devedor, ebit),
    }


@app.route("/dre")
def pagina_dre():
    from calendar import monthrange
    hoje = datetime.now(_BRT)
    try:
        mes = int(request.args.get("mes", hoje.month))
        ano = int(request.args.get("ano", hoje.year))
        if not (1 <= mes <= 12):
            mes = hoje.month
    except (ValueError, TypeError):
        mes, ano = hoje.month, hoje.year

    acumulado = request.args.get("acumulado", "0") == "1"

    # Build current period date range
    if acumulado:
        d_ini = datetime(ano, 1, 1)
        d_fim = datetime(ano, mes, monthrange(ano, mes)[1], 23, 59, 59)
    else:
        d_ini = datetime(ano, mes, 1)
        d_fim = datetime(ano, mes, monthrange(ano, mes)[1], 23, 59, 59)

    # Build previous period date range (previous month, or same period -1 year for acumulado)
    if acumulado:
        d_ini_prev = datetime(ano - 1, 1, 1)
        d_fim_prev = datetime(ano - 1, mes, monthrange(ano - 1, mes)[1], 23, 59, 59)
    else:
        prev_mes = mes - 1 if mes > 1 else 12
        prev_ano = ano if mes > 1 else ano - 1
        d_ini_prev = datetime(prev_ano, prev_mes, 1)
        d_fim_prev = datetime(prev_ano, prev_mes, monthrange(prev_ano, prev_mes)[1], 23, 59, 59)

    todos = _dre_ler_lancamentos()

    def _filtrar(d0, d1):
        return [l for l in todos if d0 <= l["dt"] <= d1]

    dre_atual = _dre_calcular(_filtrar(d_ini, d_fim))
    dre_prev  = _dre_calcular(_filtrar(d_ini_prev, d_fim_prev))

    meses_disponiveis = sorted({(l["dt"].year, l["dt"].month) for l in todos})

    return render_template("dre.html",
        active="dre",
        dre=dre_atual,
        dre_prev=dre_prev,
        mes=mes, ano=ano,
        acumulado=acumulado,
        meses_disponiveis=meses_disponiveis,
        label_atual=_nome_mes_label(mes, ano, acumulado),
        label_prev=_nome_mes_label(d_ini_prev.month, d_ini_prev.year, acumulado),
    )


@app.route("/dre/api/recalcular", methods=["POST"])
def dre_api_recalcular():
    from calendar import monthrange
    try:
        dados = request.get_json(force=True, silent=True) or {}
        lancamentos_raw = dados.get("lancamentos", [])
        mes = int(dados.get("mes", 1))
        ano = int(dados.get("ano", datetime.now(_BRT).year))
        acumulado = bool(dados.get("acumulado", False))
        if not (1 <= mes <= 12):
            return jsonify({"ok": False, "erro": "Mês inválido"}), 400

        # Parse novos lançamentos do arquivo importado
        novos = []
        for item in lancamentos_raw:
            try:
                dt = datetime.strptime(str(item.get("dt", "")), "%Y-%m-%d")
                valor = float(item.get("valor", 0))
                cod = str(item.get("codigo", "")).strip()
                num = str(item.get("num", "")).strip()
                if cod:
                    novos.append({"codigo": cod, "dt": dt, "valor": valor, "num": num})
            except Exception:
                continue

        # Mescla: novos têm prioridade; existentes preenchem o restante
        vistos = {(l["num"], l["codigo"]) for l in novos if l["num"]}
        existentes = _dre_ler_lancamentos() + _ler_lancamentos_jun_jul()
        for l in existentes:
            chave = (l.get("num", ""), l["codigo"])
            if chave[0] and chave in vistos:
                continue  # duplicata já presente no arquivo novo
            vistos.add(chave)
            novos.append(l)

        d_ini = datetime(ano, 1, 1) if acumulado else datetime(ano, mes, 1)
        d_fim = datetime(ano, mes, monthrange(ano, mes)[1], 23, 59, 59)
        if acumulado:
            d_ini_prev = datetime(ano - 1, 1, 1)
            d_fim_prev = datetime(ano - 1, mes, monthrange(ano - 1, mes)[1], 23, 59, 59)
        else:
            pm = mes - 1 if mes > 1 else 12
            py = ano if mes > 1 else ano - 1
            d_ini_prev = datetime(py, pm, 1)
            d_fim_prev = datetime(py, pm, monthrange(py, pm)[1], 23, 59, 59)

        def _f(d0, d1):
            return [l for l in novos if d0 <= l["dt"] <= d1]

        return jsonify({
            "ok": True,
            "dre":      _dre_calcular(_f(d_ini, d_fim)),
            "dre_prev": _dre_calcular(_f(d_ini_prev, d_fim_prev)),
        })
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Financiamentos & Consórcios ───────────────────────────────────────────────

@app.route("/financiamentos")
def pagina_financiamentos():
    from math import ceil
    hoje = datetime.now(_BRT).date()

    sb   = _supabase()
    rows = sb.table("financiamentos_contratos").select("*").order("created_at").execute().data or []

    def _restante(vcto_str):
        if not vcto_str:
            return 0
        try:
            vcto = date.fromisoformat(str(vcto_str)[:10])
        except Exception:
            return 0
        dias = (vcto - hoje).days
        return ceil(dias / 30.44) if dias > 0 else 0

    contratos = []
    for r in rows:
        parcelas = int(r["parcelas_total"])
        parcela  = float(r["valor_parcela"])
        restante = _restante(r.get("data_vencimento"))
        pagas    = parcelas - restante

        contratos.append({
            "operacao":        r["operacao"],
            "contrato":        r.get("contrato") or "",
            "placa":           r.get("placa") or "",
            "data_vencimento": str(r.get("data_vencimento") or "")[:10] or None,
            "restante":        restante,
            "parcelas_total":  parcelas,
            "valor_parcela":   parcela,
            "total_pago":      pagas * parcela,
            "total":           parcelas * parcela,
            "devedor":         restante * parcela,
            "c_prazo":         min(restante, 12) * parcela,
            "l_prazo":         max(restante - 12, 0) * parcela,
            "pct_quitado":     pagas / parcelas if parcelas else 0,
            "quitado":         restante == 0,
        })

    contratos.sort(key=lambda x: x["pct_quitado"], reverse=True)

    ativos   = [c for c in contratos if not c["quitado"]]
    quitados = [c for c in contratos if     c["quitado"]]

    soma_devedor    = sum(c["devedor"]       for c in ativos)
    soma_total_pago = sum(c["total_pago"]    for c in contratos)
    soma_c_prazo    = sum(c["c_prazo"]       for c in ativos)
    soma_l_prazo    = sum(c["l_prazo"]       for c in ativos)
    soma_mensal     = sum(c["valor_parcela"] for c in ativos)
    tempo_medio     = sum(c["restante"]      for c in ativos) / len(ativos) if ativos else 0

    ativos_vcto = [c for c in ativos if c["data_vencimento"]]
    mais_perto  = min(ativos_vcto, key=lambda x: x["data_vencimento"])["operacao"] if ativos_vcto else "—"

    cards = {
        "saldo_devedor": soma_devedor,
        "total_pago":    soma_total_pago,
        "curto_prazo":   soma_c_prazo,
        "longo_prazo":   soma_l_prazo,
        "valor_mensal":  soma_mensal,
        "tempo_medio":   round(tempo_medio, 1),
        "mais_perto":    mais_perto,
        "r_quitados":    sum(c["total_pago"] for c in quitados),
        "pct_cp":        soma_c_prazo / soma_devedor if soma_devedor else 0,
        "pct_lp":        soma_l_prazo / soma_devedor if soma_devedor else 0,
    }

    return render_template("financiamentos.html",
        active="financiamentos",
        contratos=contratos,
        cards=cards,
    )


# ── Capital Investido ─────────────────────────────────────────────────────────

_CI_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRZFq7H45YwN5Sbc9yZbSc9HGcTOl99X2jC2TVYVT828yCpilVXhzT55-W3Ma5ctQ"
    "/pub?gid=391372590&single=true&output=csv"
)
_CI_CACHE = {"day": None, "text": ""}


def _ci_fetch_csv():
    import requests as _req
    today = datetime.now(_BRT).strftime("%Y-%m-%d")
    if _CI_CACHE["day"] == today and _CI_CACHE["text"]:
        return _CI_CACHE["text"], None
    try:
        resp = _req.get(_CI_CSV_URL, timeout=15,
                        headers={"User-Agent": "Mozilla/5.0"},
                        allow_redirects=True)
        resp.raise_for_status()
        text = resp.content.decode("utf-8")
        _CI_CACHE["day"]  = today
        _CI_CACHE["text"] = text
        return text, None
    except Exception as e:
        return "", str(e)


def _fin_total_pago():
    from math import ceil
    try:
        sb   = _supabase()
        rows = sb.table("financiamentos_contratos").select("*").execute().data or []
        hoje = datetime.now(_BRT).date()
        total = 0.0
        for r in rows:
            parcelas = int(r["parcelas_total"])
            parcela  = float(r["valor_parcela"])
            vcto_str = r.get("data_vencimento")
            restante = 0
            if vcto_str:
                try:
                    vcto = date.fromisoformat(str(vcto_str)[:10])
                    dias = (vcto - hoje).days
                    restante = ceil(dias / 30.44) if dias > 0 else 0
                except Exception:
                    pass
            pagas = parcelas - restante
            total += pagas * parcela
        return total
    except Exception:
        return 0.0


@app.route("/capital-investido")
def pagina_capital_investido():
    csv_text, csv_error = _ci_fetch_csv()
    total_pago = _fin_total_pago()
    try:
        sb = _supabase()
        res = sb.table("capital_aportes").select(
            "data, investidor, descricao, banco_destino, valor"
        ).order("data").execute()
        aportes_extra = [
            {
                "data":          str(r["data"]),
                "investidor":    r["investidor"],
                "descricao":     r.get("descricao") or "",
                "banco_destino": r.get("banco_destino") or "",
                "valor":         float(r["valor"]),
            }
            for r in (res.data or [])
        ]
    except Exception:
        aportes_extra = []
    return render_template("capital_investido.html",
        active="capital_investido",
        csv_text=csv_text,
        csv_error=csv_error,
        total_pago=total_pago,
        aportes_extra=aportes_extra,
    )


@app.route("/api/capital/aportes", methods=["POST"])
def api_capital_aportes():
    body       = request.get_json(silent=True) or {}
    data       = (body.get("data")       or "").strip()
    investidor = (body.get("investidor") or "").strip()
    descricao  = (body.get("descricao")  or "").strip()
    banco_dest = (body.get("banco_destino") or "").strip()
    try:
        valor = float(str(body.get("valor", "")).replace(",", "."))
    except (ValueError, TypeError):
        return jsonify({"ok": False, "erro": "Valor inválido"}), 400
    if not data or not investidor:
        return jsonify({"ok": False, "erro": "Data e investidor são obrigatórios"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 500
    sb.table("capital_aportes").insert({
        "data":          data,
        "investidor":    investidor,
        "descricao":     descricao,
        "banco_destino": banco_dest,
        "valor":         valor,
    }).execute()
    return jsonify({"ok": True, "data": data, "investidor": investidor, "valor": valor})


# ── Frota ─────────────────────────────────────────────────────────────────────

def _ler_frota_dados():
    """
    Lê veículos e histórico FIPE do Supabase (frota_veiculos + frota_fipe_historico).
    Retorna (veiculos, codigos, erro) com a mesma estrutura anterior para compatibilidade
    com o template frota.html.
    """
    try:
        sb = _supabase()
        if sb is None:
            return [], [], "Supabase não configurado."

        res_v = sb.table("frota_veiculos").select(
            "modelo, placa, ano_modelo, cod_fipe, dt_aquisicao, vl_aquisicao"
        ).eq("ativo", True).execute()

        res_h = sb.table("frota_fipe_historico").select(
            "placa, mes_ref, valor"
        ).eq("fonte", "planilha").execute()

        # Pivot histórico: {placa: {python_key: valor}}
        # 'JAN/25' → 'jan25', 'MAI/26' → 'mai26'
        hist: dict[str, dict[str, float]] = {}
        for row in (res_h.data or []):
            key = row["mes_ref"].replace("/", "").lower()
            hist.setdefault(row["placa"], {})[key] = float(row["valor"])

        _EXCLUIR_PLACAS = {"QGO-2H58"}
        veiculos = []
        for v in (res_v.data or []):
            placa = v["placa"]
            if placa.upper() in _EXCLUIR_PLACAS:
                continue
            meses = hist.get(placa, {})
            dt = v.get("dt_aquisicao") or ""
            try:
                dt = datetime.strptime(dt, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass
            veiculos.append({
                "modelo":       v["modelo"],
                "placa":        placa,
                "ano_modelo":   v["ano_modelo"],
                "cod_fipe":     v["cod_fipe"],
                "dt_aquisicao": dt,
                "vl_aquisicao": float(v["vl_aquisicao"]) if v["vl_aquisicao"] is not None else None,
                **{k: meses.get(k) for k in (
                    "jan25","fev25","mar25","abr25","mai25","jun25",
                    "jul25","ago25","set25","out25","nov25","dez25",
                    "jan26","fev26","mar26","abr26","mai26","jun26",
                    "jul26","ago26","set26","out26","nov26","dez26",
                )},
            })

        # Derive codigos from vehicles (used by template for CODIGOS JS var)
        seen: dict[tuple, dict] = {}
        for v in veiculos:
            key = (v["cod_fipe"], v["ano_modelo"])
            if key not in seen:
                seen[key] = {"cod_fipe": v["cod_fipe"], "modelo": v["modelo"],
                             "ano_modelo": v["ano_modelo"], "qtd": 0}
            seen[key]["qtd"] += 1
        codigos = list(seen.values())

        return veiculos, codigos, None
    except Exception as e:
        import traceback; traceback.print_exc()
        return [], [], str(e)


def _frota_mes_atual():
    """Retorna (curr_key, curr_label, prev_key, prev_label) baseado na data do sistema."""
    MESES = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
    hoje  = datetime.now(_BRT)
    ano, mes = hoje.year, hoje.month
    yy        = str(ano)[2:]
    curr_key   = MESES[mes - 1] + yy
    curr_label = MESES[mes - 1].upper() + '/' + yy
    prev_mes  = mes - 1 if mes > 1 else 12
    prev_ano  = ano if mes > 1 else ano - 1
    prev_yy   = str(prev_ano)[2:]
    prev_key   = MESES[prev_mes - 1] + prev_yy
    prev_label = MESES[prev_mes - 1].upper() + '/' + prev_yy
    return curr_key, curr_label, prev_key, prev_label


def _frota_ler_manual():
    """Retorna {placa: {mes_ref: {valor, atualizado_em}}} lido de frota_fipe_historico."""
    try:
        sb = _supabase()
        if sb is None:
            return {}
        res = sb.table("frota_fipe_historico").select(
            "placa, mes_ref, valor, atualizado_em"
        ).eq("fonte", "manual").execute()
        out = {}
        for row in (res.data or []):
            placa  = row["placa"]
            ref    = row["mes_ref"]
            dt_str = row.get("atualizado_em") or ""
            try:
                dt_str = datetime.strptime(dt_str, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass
            out.setdefault(placa, {})[ref] = {
                "valor":         float(row["valor"]),
                "atualizado_em": dt_str,
            }
        return out
    except Exception:
        import traceback; traceback.print_exc()
        return {}


def _frota_salvar_manual(placa, valor, ref):
    """Upsert de (placa, mes_ref) em frota_fipe_historico com fonte='manual'."""
    sb = _supabase()
    if sb is None:
        return
    sb.table("frota_fipe_historico").upsert({
        "placa":         placa,
        "mes_ref":       ref,
        "valor":         valor,
        "fonte":         "manual",
        "atualizado_em": datetime.now(_BRT).strftime("%Y-%m-%d"),
    }, on_conflict="placa,mes_ref").execute()


_SOB_ADM_FIPE_VALORES = {
    "005540-9": 800.0,
    "095010-6": 1200.0,
}
_SOB_ADM_PLACA_VALORES = {}
_SOB_ADM_TAXA = 0.15
_SOB_ADM_PLACA_EXTRA = ""


def _ler_sob_administracao():
    """Lê DADOS_CLIENTES_CONS.xlsx: veículos cuja unidade não seja Ativuz/AZ ou placa especial."""
    import openpyxl
    xlsx_path = _clientes_cons_xlsx_path()
    if not xlsx_path.exists():
        return [], None

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return [], None

        header = rows[0]

        def _norm(s):
            s = unicodedata.normalize("NFD", str(s or "").lower())
            return "".join(c for c in s if unicodedata.category(c) != "Mn")

        def _ci(kw):
            nk = _norm(kw)
            return next((i for i, h in enumerate(header) if nk in _norm(str(h or ""))), None)

        def _ci_exact(name):
            n = _norm(name)
            return next((i for i, h in enumerate(header) if _norm(str(h or "")) == n), None)

        i_placa    = _ci("placa")
        i_mod      = _ci_exact("modelo") if _ci_exact("modelo") is not None else _ci("modelo")
        i_marca    = _ci("marca")
        i_prop     = _ci("unidade do ve")
        i_loc      = _ci("razao social cliente") or _ci("razao social") or _ci("cliente")
        i_tipo     = _ci("tipo de contrato")
        i_ini      = _ci("inicio do contrato") or _ci("inicio de contrato") or _ci("inicio")
        i_fim      = _ci("termino do contrato") or _ci("termino")
        i_anomod   = _ci("ano modelo")

        def _v(row, i):
            if i is None or i >= len(row): return ""
            v = row[i]
            if v is None: return ""
            if hasattr(v, "strftime"): return v.strftime("%d/%m/%Y")
            return str(v).strip()

        _EXTRA = _SOB_ADM_PLACA_EXTRA.upper().replace("-", "").replace(" ", "")

        hoje = date.today()
        veiculos = []
        for row in rows[1:]:
            placa = _v(row, i_placa)
            if not placa:
                continue

            placa_id = placa.upper().replace("-", "").replace(" ", "")
            marca    = _norm(_v(row, i_marca))
            modelo   = _norm(_v(row, i_mod))

            # inclui: Polo, BYD ou placa especial
            eh_polo  = "polo" in modelo
            eh_byd   = "byd" in marca
            eh_extra = bool(_EXTRA) and placa_id == _EXTRA
            if not (eh_polo or eh_byd or eh_extra):
                continue

            unid = "GC AUTOELÉTRICA" if eh_extra else _v(row, i_prop)

            if eh_polo:
                valor_s = 800.0
            elif eh_byd:
                valor_s = 1200.0
            else:
                valor_s = _SOB_ADM_PLACA_VALORES.get(placa.upper())
            taxa_s  = round(valor_s * _SOB_ADM_TAXA, 2) if valor_s else None

            ini_raw = row[i_ini] if i_ini is not None and i_ini < len(row) else None
            ini_date = None
            if ini_raw:
                if isinstance(ini_raw, datetime): ini_date = ini_raw.date()
                elif isinstance(ini_raw, date):   ini_date = ini_raw
                else:
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                        try: ini_date = datetime.strptime(str(ini_raw)[:10], fmt).date(); break
                        except ValueError: pass

            dias_ativos = (hoje - ini_date).days if ini_date else 0
            receita_acum = round(dias_ativos * (taxa_s / 7), 2) if taxa_s and dias_ativos > 0 else 0.0

            ini_fmt = ini_date.strftime("%d/%m/%Y") if ini_date else ""

            veiculos.append({
                "placa":        placa,
                "montadora":    _v(row, i_marca),
                "modelo":       _v(row, i_mod),
                "fipe":         "",
                "ano_fab":      _v(row, i_anomod),
                "ano_mod":      _v(row, i_anomod),
                "proprietario": unid,
                "locatario":    _v(row, i_loc),
                "tipo_contrato":_v(row, i_tipo),
                "km":           "",
                "inicio":       ini_fmt,
                "termino":      _v(row, i_fim),
                "situacao":     "EM ANDAMENTO",
                "valor_semanal":valor_s,
                "taxa_semanal": taxa_s,
                "dias_ativos":  dias_ativos,
                "receita_acum": receita_acum,
            })

        veiculos.sort(key=lambda v: (v["proprietario"], v["inicio"]))
        return veiculos, None

    except Exception as e:
        import traceback; traceback.print_exc()
        return [], str(e)


def _gerar_segundas(ini, fim):
    """Retorna todas as segundas-feiras em [ini, fim] inclusive."""
    from datetime import timedelta
    days = (7 - ini.weekday()) % 7   # 0 se já é segunda
    cur = ini + timedelta(days=days)
    result = []
    while cur <= fim:
        result.append(cur)
        cur += timedelta(weeks=1)
    return result


@app.route("/api/sob-adm/recebimentos")
def api_sob_adm_recebimentos():
    """
    Retorna cada veículo com seu calendário de segundas-feiras (passadas + futuras)
    e o status de recebimento de cada uma.

    Supabase — tabela necessária (execute uma vez):
      CREATE TABLE sob_adm_recebimentos (
        id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
        placa text NOT NULL,
        data_semana date NOT NULL,
        taxa_valor numeric(10,2) NOT NULL DEFAULT 0,
        recebido boolean DEFAULT false,
        created_at timestamptz DEFAULT now(),
        UNIQUE(placa, data_semana)
      );
    """
    from datetime import timedelta
    sob_adm, erro = _ler_sob_administracao()
    if erro:
        return jsonify({"ok": False, "erro": erro})

    # Carrega registros do Supabase
    recebidos = {}   # (placa, "YYYY-MM-DD") -> bool
    sb = _supabase()
    if sb:
        try:
            res = sb.table("sob_adm_recebimentos").select("placa,data_semana,recebido").execute()
            for r in (res.data or []):
                ds = str(r["data_semana"])[:10]
                recebidos[(r["placa"], ds)] = bool(r.get("recebido", False))
        except Exception:
            pass

    hoje = date.today()
    result = []

    for v in sob_adm:
        if not v["inicio"] or not v["taxa_semanal"]:
            continue
        try:
            ini = datetime.strptime(v["inicio"], "%d/%m/%Y").date()
        except ValueError:
            continue

        fim_contrato = None
        if v.get("termino"):
            for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                try:
                    fim_contrato = datetime.strptime(str(v["termino"])[:10], fmt).date()
                    break
                except ValueError:
                    pass

        ate_hoje = min(hoje, fim_contrato) if fim_contrato else hoje
        segundas_passadas = _gerar_segundas(ini, ate_hoje)
        segundas_futuras  = _gerar_segundas(hoje + timedelta(days=1), fim_contrato) if fim_contrato and fim_contrato > hoje else []

        semanas = []
        for d in segundas_passadas:
            ds = d.isoformat()
            semanas.append({"data": ds, "recebido": recebidos.get((v["placa"], ds), False), "passada": True})
        for d in segundas_futuras:
            ds = d.isoformat()
            semanas.append({"data": ds, "recebido": False, "passada": False})

        taxa = v["taxa_semanal"]
        rec  = round(sum(taxa for s in semanas if s["passada"] and s["recebido"]), 2)
        pend = round(sum(taxa for s in semanas if s["passada"] and not s["recebido"]), 2)
        fut  = round(sum(taxa for s in semanas if not s["passada"]), 2)

        result.append({
            "placa":            v["placa"],
            "modelo":           v["modelo"],
            "proprietario":     v["proprietario"],
            "locatario":        v["locatario"],
            "valor_semanal":    v["valor_semanal"],
            "taxa_semanal":     taxa,
            "inicio":           v["inicio"],
            "termino":          v.get("termino") or None,
            "semanas":          semanas,
            "recebido_total":   rec,
            "pendente_total":   pend,
            "projetado_futuro": fut,
        })

    return jsonify({"ok": True, "veiculos": result})


@app.route("/api/sob-adm/recebimentos/toggle", methods=["POST"])
def api_sob_adm_toggle():
    dados = request.get_json(force=True, silent=True) or {}
    placa       = str(dados.get("placa", "")).strip()
    data_semana = str(dados.get("data_semana", "")).strip()

    if not placa or not data_semana:
        return jsonify({"ok": False, "erro": "Parâmetros inválidos"}), 400

    sob_adm, _ = _ler_sob_administracao()
    v = next((x for x in sob_adm if x["placa"] == placa), None)
    taxa = float(v["taxa_semanal"]) if v and v.get("taxa_semanal") else 0.0

    sb = _supabase()
    if not sb:
        return jsonify({"ok": False, "erro": "Banco indisponível"}), 503

    try:
        res = sb.table("sob_adm_recebimentos").select("id,recebido").eq("placa", placa).eq("data_semana", data_semana).execute()
        if res.data:
            novo = not res.data[0]["recebido"]
            sb.table("sob_adm_recebimentos").update({"recebido": novo}).eq("id", res.data[0]["id"]).execute()
        else:
            novo = True
            sb.table("sob_adm_recebimentos").insert({"placa": placa, "data_semana": data_semana, "taxa_valor": taxa, "recebido": True}).execute()
        return jsonify({"ok": True, "recebido": novo})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500



_CONTRATOS_XLSX = Path(__file__).parent / "planilhas" / "Contratos de Locação.xlsx"


def _ler_contratos():
    import openpyxl
    if not _CONTRATOS_XLSX.exists():
        return [], f"Arquivo não encontrado: {_CONTRATOS_XLSX.name}"
    try:
        wb = openpyxl.load_workbook(str(_CONTRATOS_XLSX), data_only=True)
        ws = wb["Relatório"]
        today = date.today()

        def _fmt(v):
            if isinstance(v, datetime): return v.strftime('%d/%m/%Y')
            if isinstance(v, date):     return v.strftime('%d/%m/%Y')
            return str(v) if v else ''

        rows = []
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
            if not row[9]:  # Contrato Comercial
                continue
            termino_raw = row[52]  # Término Previsto
            dias_vencer = None
            if isinstance(termino_raw, datetime):
                dias_vencer = (termino_raw.date() - today).days
            rows.append({
                'contrato_comercial': str(row[9]  or ''),   # Contrato Comercial
                'contrato_locacao':   str(row[12] or ''),   # Contrato de Locação
                'periodo':            str(row[41] or ''),   # Período (meses)
                'cliente':            str(row[6]  or ''),   # Cliente
                'unidade_fat':        str(row[54] or ''),   # Unidade de faturamento
                'valor_locacao':      float(row[57] or 0),  # Valor de locação vigente
                'tipo_pessoa':        str(row[51] or ''),   # Tipo de cliente
                'gasto_total':        float(row[27] or 0),  # Gasto Total
                'gasto_sinistros':    float(row[26] or 0),  # Gasto Sinistros
                'gasto_manutencao':   float(row[25] or 0),  # Gasto Manutenção
                'inicio':             _fmt(row[30]),         # Início de Contrato
                'termino_previsto':   _fmt(row[52]),         # Término Previsto
                'situacao':           str(row[46] or ''),   # Situação
                'placa':              str(row[59] or ''),   # Veículo Atual
                'modelo':             str(row[36] or ''),   # Modelo
                'km':                 int(row[31]  or 0),   # Km confirmado
                'grupo':              str(row[28] or ''),   # Grupo
                'tipo_contrato':      str(row[50] or ''),   # Tipo de Contrato
                'sit_faturamento':    str(row[47] or ''),   # Situação de Faturamento
                'valor_inicial':      float(row[58] or 0),  # Valor inicial de locação
                'dias_vencer':        dias_vencer,
            })
        wb.close()
        return rows, None
    except Exception as ex:
        return [], str(ex)


@app.route("/insights/contratos")
def pagina_contratos():
    contratos, erro = _ler_contratos()
    ativos = [c for c in contratos if c['situacao'] == 'EM ANDAMENTO']

    receita_mes    = sum(c['valor_locacao'] for c in ativos)
    gasto_acum     = sum(c['gasto_total']   for c in contratos)
    a_vencer_30    = sum(1 for c in ativos
                         if c['dias_vencer'] is not None and 0 <= c['dias_vencer'] <= 30)

    tipos_count = {}
    for c in contratos:
        t = c['tipo_contrato'] or 'Não informado'
        tipos_count[t] = tipos_count.get(t, 0) + 1

    clientes_rec = {}
    for c in ativos:
        clientes_rec[c['cliente']] = clientes_rec.get(c['cliente'], 0) + c['valor_locacao']
    top10 = sorted(clientes_rec.items(), key=lambda x: -x[1])[:10]

    return render_template("contratos.html",
        active="contratos",
        contratos=contratos,
        erro=erro,
        kpi_ativos=len(ativos),
        kpi_receita=receita_mes,
        kpi_vencer30=a_vencer_30,
        kpi_gasto=gasto_acum,
        tipos_count=tipos_count,
        top10_clientes=top10,
    )


@app.route("/insights/frota")
def pagina_frota():
    veiculos, codigos, erro = _ler_frota_dados()
    manual = _frota_ler_manual()
    curr_key, curr_label, prev_key, prev_label = _frota_mes_atual()
    sob_adm, sob_adm_erro = _ler_sob_administracao()
    return render_template("frota.html",
        active="frota",
        veiculos=veiculos,
        codigos=codigos,
        manual=manual,
        erro=erro,
        curr_mes_key=curr_key,
        curr_mes_label=curr_label,
        prev_mes_key=prev_key,
        prev_mes_label=prev_label,
        sob_adm=sob_adm,
        sob_adm_erro=sob_adm_erro,
    )


@app.route("/api/frota/manual", methods=["POST"])
def api_frota_manual():
    body  = request.get_json(silent=True) or {}
    placa = (body.get("placa") or "").strip().upper()
    ref   = (body.get("ref")   or "").strip()
    try:
        valor = float(str(body.get("valor", "")).replace(",", "."))
    except (ValueError, TypeError):
        return jsonify({"ok": False, "erro": "Valor inválido"}), 400
    if not placa:
        return jsonify({"ok": False, "erro": "Placa obrigatória"}), 400
    _frota_salvar_manual(placa, valor, ref)
    return jsonify({"ok": True, "placa": placa, "valor": valor, "ref": ref})


@app.route("/api/frota/manual/batch", methods=["POST"])
def api_frota_manual_batch():
    """Upsert em bulk por combinação (cod_fipe + ano_modelo) em frota_fipe_historico."""
    body     = request.get_json(silent=True) or {}
    entradas = body.get("entradas") or []
    if not entradas:
        return jsonify({"ok": False, "erro": "Nenhuma entrada"}), 400

    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 500

    res_v = sb.table("frota_veiculos").select(
        "placa, cod_fipe, ano_modelo"
    ).eq("ativo", True).execute()
    veiculos = res_v.data or []
    agora = datetime.now(_BRT).strftime("%Y-%m-%d")

    rows = []
    for entrada in entradas:
        cod     = (entrada.get("cod_fipe")   or "").strip()
        ano_mod = (entrada.get("ano_modelo") or "").strip()
        ref     = (entrada.get("ref")        or "").strip()
        try:
            valor = float(str(entrada.get("valor", "")).replace(",", "."))
        except (ValueError, TypeError):
            continue
        for v in veiculos:
            if v["cod_fipe"] == cod and ((not ano_mod) or v["ano_modelo"] == ano_mod):
                rows.append({
                    "placa":         v["placa"],
                    "mes_ref":       ref,
                    "valor":         valor,
                    "fonte":         "manual",
                    "atualizado_em": agora,
                })

    if rows:
        sb.table("frota_fipe_historico").upsert(rows, on_conflict="placa,mes_ref").execute()

    return jsonify({"ok": True, "atualizados": len(rows)})


# ── Carteira Judicializada ────────────────────────────────────────────────────

@app.route("/api/carteira-judicializada", methods=["GET"])
def api_carteira_judicializada_listar():
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").select("*").order("criado_em").execute()
        return jsonify({"ok": True, "data": res.data or []})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/carteira-judicializada", methods=["POST"])
def api_carteira_judicializada_inserir():
    body = request.get_json(silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    if not cliente:
        return jsonify({"ok": False, "erro": "Cliente obrigatório"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").insert({
            "cliente":          cliente,
            "cpf_cnpj":         (body.get("cpf_cnpj")        or "").strip(),
            "avalista":         (body.get("avalista")         or "").strip(),
            "cpf_avalista":     (body.get("cpf_avalista")     or "").strip(),
            "inicio_divida":    body.get("inicio_divida")     or None,
            "valor_atual":      float(body.get("valor_atual") or 0),
            "status":           body.get("status")            or "Ajuizado",
            "num_processo":     (body.get("num_processo")     or "").strip(),
            "proximo_prazo":    body.get("proximo_prazo")     or None,
            "descricao_prazo":  (body.get("descricao_prazo")  or "").strip(),
        }).execute()
        return jsonify({"ok": True, "data": res.data[0] if res.data else {}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/carteira-judicializada/<uuid:registro_id>", methods=["PUT"])
def api_carteira_judicializada_atualizar(registro_id):
    body = request.get_json(silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    if not cliente:
        return jsonify({"ok": False, "erro": "Cliente obrigatório"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("carteira_judicializada").update({
            "cliente":          cliente,
            "cpf_cnpj":         (body.get("cpf_cnpj")        or "").strip(),
            "avalista":         (body.get("avalista")         or "").strip(),
            "cpf_avalista":     (body.get("cpf_avalista")     or "").strip(),
            "inicio_divida":    body.get("inicio_divida")     or None,
            "valor_atual":      float(body.get("valor_atual") or 0),
            "status":           body.get("status")            or "Ajuizado",
            "num_processo":     (body.get("num_processo")     or "").strip(),
            "proximo_prazo":    body.get("proximo_prazo")     or None,
            "descricao_prazo":  (body.get("descricao_prazo")  or "").strip(),
            "atualizado_em":    "now()",
        }).eq("id", str(registro_id)).execute()
        return jsonify({"ok": True, "data": res.data[0] if res.data else {}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Acordo Judicializado ─────────────────────────────────────────────────────

@app.route("/api/carteira-judicializada/<uuid:registro_id>/acordo", methods=["PUT"])
def api_jud_acordo_salvar(registro_id):
    body = request.get_json(silent=True) or {}
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("carteira_judicializada").update({
            "acordo_dados": body.get("acordo_dados"),
        }).eq("id", str(registro_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Checklist Judicializada ───────────────────────────────────────────────────

@app.route("/api/jud-checklist/<uuid:registro_id>", methods=["GET"])
def api_jud_checklist_get(registro_id):
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        key = "JUD-" + str(registro_id)
        res = sb.table("checklist_contratos").select("*").eq("contrato", key).execute()
        if res.data:
            contrato_id = res.data[0]["id"]
        else:
            ins = sb.table("checklist_contratos").insert({"contrato": key, "placa": "", "cliente": "", "unidade": ""}).execute()
            contrato_id = ins.data[0]["id"]
        itens = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute().data or []
        return jsonify({"ok": True, "contrato_id": contrato_id,
                        "itens": [{"id": i["id"], "nome": i["nome"], "marcado": i["marcado"]} for i in itens]})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/jud-checklist/<uuid:registro_id>/item", methods=["POST"])
def api_jud_checklist_add_item(registro_id):
    body = request.get_json(silent=True) or {}
    contrato_id = body.get("contrato_id")
    nome = (body.get("nome") or "").strip()
    if not contrato_id or not nome:
        return jsonify({"ok": False, "erro": "contrato_id e nome são obrigatórios"}), 400
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        res = sb.table("checklist_itens").insert({"contrato_id": contrato_id, "nome": nome, "marcado": False}).execute()
        item = res.data[0]
        return jsonify({"ok": True, "item": {"id": item["id"], "nome": item["nome"], "marcado": item["marcado"]}})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/jud-checklist/item/<uuid:item_id>", methods=["PUT"])
def api_jud_checklist_toggle(item_id):
    body = request.get_json(silent=True) or {}
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("checklist_itens").update({"marcado": bool(body.get("marcado", False))}).eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/jud-checklist/item/<uuid:item_id>", methods=["DELETE"])
def api_jud_checklist_delete(item_id):
    sb = _supabase()
    if sb is None:
        return jsonify({"ok": False, "erro": "Supabase não configurado"}), 503
    try:
        sb.table("checklist_itens").delete().eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


# ── Checklist ─────────────────────────────────────────────────────────────────

def _clientes_cons_xlsx_path():
    base = Path(__file__).resolve().parent
    return base / "planilhas" / "DADOS_CLIENTES_CONS.xlsx"


_IMAGEM_MAP = [
    ("GOL",     "gol_sf.png"),
    ("VOYAGE",  "voyage.png"),
    ("POLO",    "polo.png"),
    ("DOLPHIN", "byd.png"),
    ("SANDERO", "sandero.png"),
    ("CG",      "CG.png"),
    ("NXR",     "nxr.png"),
]
_BLEND_MULTIPLY = {"gol_sf.png", "voyage.png", "polo.png", "byd.png", "CG.png"}


def _imagem_veiculo(modelo):
    m = (modelo or "").upper()
    for keyword, fname in _IMAGEM_MAP:
        if keyword in m:
            return fname
    return None


def _ler_veiculos():
    import openpyxl
    xlsx_path = _clientes_cons_xlsx_path()
    if not xlsx_path.exists():
        return [], "Planilha não encontrada em planilhas/DADOS_CLIENTES_CONS.xlsx."

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 1:
            return [], "Planilha sem dados."

        header_row = rows[0]

        def _norm(s):
            s = unicodedata.normalize("NFD", str(s or "").lower())
            return "".join(c for c in s if unicodedata.category(c) != "Mn")

        headers_norm = [_norm(h) for h in header_row]

        def _ci(keyword):
            kn = _norm(keyword)
            return next((i for i, h in enumerate(headers_norm) if kn in h), None)

        def _ci_exact(name):
            n = _norm(name)
            return next((i for i, h in enumerate(headers_norm) if h == n), None)

        i_placa    = _ci("placa")
        i_modelo   = _ci_exact("modelo") if _ci_exact("modelo") is not None else _ci("modelo")
        i_cliente  = _ci("razao social cliente") or _ci("razao social") or _ci("cliente")
        i_contrato = _ci("contrato de locacao") or _ci("contrato")
        i_unidade  = _ci("unidade do veiculo") or _ci("unidade")
        i_inicio   = _ci("inicio de contrato") or _ci("inicio")
        i_termino  = _ci("termino de contrato") or _ci("termino")

        def _v(row, i):
            if i is None or i >= len(row):
                return ""
            v = row[i]
            if v is None:
                return ""
            if hasattr(v, "strftime"):
                return v.strftime("%d/%m/%Y")
            return str(v).strip()

        seen_placas = {}
        for row in rows[1:]:
            placa = _v(row, i_placa)
            if not placa:
                continue
            modelo = _v(row, i_modelo)
            img    = _imagem_veiculo(modelo)
            seen_placas[placa] = {
                "placa":    placa,
                "modelo":   modelo,
                "cliente":  _v(row, i_cliente),
                "contrato": _v(row, i_contrato),
                "unidade":  _v(row, i_unidade),
                "inicio":   _v(row, i_inicio),
                "termino":  _v(row, i_termino),
                "imagem":   img,
                "blend":    img in _BLEND_MULTIPLY if img else False,
            }
        veiculos = sorted(seen_placas.values(), key=lambda v: v["cliente"].lower())
        return veiculos, None

    except Exception as e:
        import traceback; traceback.print_exc()
        return [], str(e)


@app.route("/checklist")
def pagina_checklist():
    veiculos, erro_leitura = _ler_veiculos()

    badge_data = {}
    sb = _supabase()
    if sb:
        try:
            contratos_res = sb.table("checklist_contratos").select("id, contrato").execute()
            if contratos_res.data:
                ids_map = {r["id"]: r["contrato"] for r in contratos_res.data}
                itens_res = sb.table("checklist_itens").select("contrato_id, marcado").execute()
                for item in (itens_res.data or []):
                    cid  = item["contrato_id"]
                    cnum = ids_map.get(cid)
                    if cnum:
                        if cnum not in badge_data:
                            badge_data[cnum] = {"total": 0, "marcados": 0}
                        badge_data[cnum]["total"] += 1
                        if item["marcado"]:
                            badge_data[cnum]["marcados"] += 1
        except Exception:
            pass

    return render_template("checklist.html",
                           active="checklist",
                           veiculos=veiculos,
                           badge_data=badge_data,
                           erro_leitura=erro_leitura)


@app.route("/api/checklist/contrato")
def api_checklist_get():
    contrato = request.args.get("contrato", "").strip()
    placa    = request.args.get("placa", "").strip()
    cliente  = request.args.get("cliente", "").strip()
    unidade  = request.args.get("unidade", "").strip()

    if not contrato:
        return jsonify({"error": "Contrato obrigatório"}), 400

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    ITENS_PADRAO = ["INDICAÇÃO DE CONDUTOR", "CONTRATO", "NOTA PROMISSÓRIA", "CHAVE RESERVA"]

    res = sb.table("checklist_contratos").select("*").eq("contrato", contrato).execute()

    if res.data:
        rec        = res.data[0]
        contrato_id = rec["id"]
        itens_res  = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute()
        itens = itens_res.data or []
        if not itens:
            for nome in ITENS_PADRAO:
                sb.table("checklist_itens").insert({"contrato_id": contrato_id, "nome": nome, "marcado": False}).execute()
            itens = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute().data or []
    else:
        ins = sb.table("checklist_contratos").insert({
            "contrato": contrato, "placa": placa, "cliente": cliente, "unidade": unidade,
        }).execute()
        rec         = ins.data[0]
        contrato_id = rec["id"]
        for nome in ITENS_PADRAO:
            sb.table("checklist_itens").insert({"contrato_id": contrato_id, "nome": nome, "marcado": False}).execute()
        itens = sb.table("checklist_itens").select("*").eq("contrato_id", contrato_id).order("created_at").execute().data or []

    return jsonify({
        "contrato_id": contrato_id,
        "contrato":    contrato,
        "placa":       rec.get("placa", placa),
        "cliente":     rec.get("cliente", cliente),
        "unidade":     rec.get("unidade", unidade),
        "itens": [{"id": i["id"], "nome": i["nome"], "marcado": i["marcado"]} for i in itens],
    })


@app.route("/api/checklist/salvar", methods=["POST"])
def api_checklist_salvar():
    data        = request.get_json(force=True)
    contrato_id = data.get("contrato_id")
    itens       = data.get("itens", [])

    if not contrato_id:
        return jsonify({"error": "contrato_id obrigatório"}), 400

    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503

    try:
        for item in itens:
            if item.get("is_new"):
                sb.table("checklist_itens").insert({
                    "contrato_id": contrato_id,
                    "nome":        item["nome"],
                    "marcado":     item.get("marcado", False),
                }).execute()
            else:
                sb.table("checklist_itens").update({
                    "marcado": bool(item.get("marcado", False))
                }).eq("id", item["id"]).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/inad/obs", methods=["PUT"])
def api_inad_obs_salvar():
    data  = request.get_json(force=True)
    chave = (data.get("chave") or "").strip()
    texto = (data.get("texto") or "").strip()[:500]
    if not chave:
        return jsonify({"error": "chave obrigatória"}), 400
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        existing = sb.table("inad_observacoes").select("id").eq("chave", chave).execute()
        if existing.data:
            sb.table("inad_observacoes").update({"texto": texto}).eq("chave", chave).execute()
        else:
            sb.table("inad_observacoes").insert({"chave": chave, "texto": texto}).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/checklist/item/<uuid:item_id>", methods=["PUT"])
def api_checklist_toggle(item_id):
    data    = request.get_json(force=True)
    marcado = bool(data.get("marcado", False))
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        sb.table("checklist_itens").update({"marcado": marcado}).eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/checklist/item", methods=["POST"])
def api_checklist_add_item():
    data        = request.get_json(force=True)
    contrato_id = data.get("contrato_id")
    nome        = (data.get("nome") or "").strip().upper()[:80]
    if not contrato_id or not nome:
        return jsonify({"error": "contrato_id e nome obrigatórios"}), 400
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        res  = sb.table("checklist_itens").insert({
            "contrato_id": contrato_id, "nome": nome, "marcado": False
        }).execute()
        item = res.data[0]
        return jsonify({"ok": True, "id": item["id"], "nome": item["nome"], "marcado": item["marcado"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/checklist/item/<uuid:item_id>", methods=["DELETE"])
def api_checklist_delete_item(item_id):
    sb = _supabase()
    if not sb:
        return jsonify({"error": "Supabase indisponível"}), 503
    try:
        sb.table("checklist_itens").delete().eq("id", str(item_id)).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Benchmarking ──────────────────────────────────────────────────────────────

@app.route("/benchmarking")
def pagina_benchmarking():
    concorrentes = benchmarking_scraper.obter_dados()
    ativuz = _calcular_indicadores_ativuz() or {"ticker": "ATIVUZ", "nome": "Ativuz",
        "erro": "Sem dados", "is_ativuz": True,
        **{k: "N/D" for k in ["pl","pvp","roe","margem_bruta","margem_ebitda",
                               "margem_ebit","margem_liquida","div_ebitda","div_ebit"]}}
    dados = [ativuz] + concorrentes
    atualizado_em = benchmarking_scraper.cache_info()

    dep_veiculos = 0
    dep_total = 0.0
    try:
        sb = _supabase()
        if sb:
            res = sb.table("frota_veiculos").select("vl_aquisicao").eq("ativo", True).execute()
            rows = res.data or []
            dep_veiculos = len(rows)
            dep_total = sum(float(r.get("vl_aquisicao") or 0) for r in rows)
    except Exception:
        pass

    def _fmt_brl(v):
        return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    return render_template("benchmarking.html", active="benchmarking",
                           dados=dados, atualizado_em=atualizado_em,
                           indicadores=benchmarking_scraper.INDICADORES,
                           depreciacao_frota_veiculos=dep_veiculos,
                           depreciacao_frota_total=_fmt_brl(dep_total),
                           depreciacao_frota_anual=_fmt_brl(dep_total / 5) if dep_total else "0,00")


@app.route("/benchmarking/atualizar", methods=["POST"])
def atualizar_benchmarking():
    benchmarking_scraper.obter_dados(forcar=True)
    return jsonify({"ok": True, "atualizado_em": benchmarking_scraper.cache_info()})


@app.route("/configuracoes")
def pagina_configuracoes():
    return render_template("configuracoes.html", active="configuracoes")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=False)
